import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.auth import User

ROLE_COLORS = {
    "ADMIN":  "1E3A5F",
    "HO":     "2563EB",
    "DM":     "059669",
    "STORE":  "D97706",
}
DEPT_COLORS = {
    "SUPERUSER":    "7C3AED",
    "IT":           "0EA5E9",
    "MANAGER":      "0D9488",
    "HR":           "EC4899",
    "FINANCE":      "F59E0B",
    "MARKETING":    "EF4444",
    "COMMERCIAL":   "8B5CF6",
    "RETAIL":       "10B981",
    "FACILITIES":   "6B7280",
    "DM":           "059669",
    "STORE":        "D97706",
    "STOREMANAGER": "B45309",
    "TOPMGR":       "1E3A5F",
}

async def main():
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(User).where(User.is_active == True)
            .order_by(User.role, User.department, User.full_name)
        )
        users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Utenti FTC HUB"

    headers = ["Username", "Nome completo", "Email", "Role", "Department", "Password", "Attivo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill("solid", fgColor="F1F5F9")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for i, u in enumerate(users, 2):
        role_val = str(u.role.value if hasattr(u.role, "value") else u.role)
        dept_val = str(u.department.value if hasattr(u.department, "value") else u.department)
        row = [
            u.username,
            u.full_name or "",
            u.email or "",
            role_val,
            dept_val,
            "123",
            "Sì" if u.is_active else "No",
        ]
        ws.append(row)
        fill = fill_even if i % 2 == 0 else fill_odd
        for cell in ws[i]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        # Colore Role
        ws.cell(row=i, column=4).font = Font(bold=True, color=ROLE_COLORS.get(role_val, "6B7280"))
        # Colore Department
        ws.cell(row=i, column=5).font = Font(bold=True, color=DEPT_COLORS.get(dept_val, "6B7280"))
        # Password in grigio corsivo
        ws.cell(row=i, column=6).font = Font(italic=True, color="6B7280")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8
    ws.freeze_panes = "A2"

    # Foglio Riepilogo
    ws2 = wb.create_sheet("Riepilogo")
    ws2.append(["Role", "Department", "Conteggio"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    from collections import Counter
    counts = Counter(
        (str(u.role.value if hasattr(u.role, "value") else u.role),
         str(u.department.value if hasattr(u.department, "value") else u.department))
        for u in users
    )
    for (role, dept), count in sorted(counts.items()):
        ws2.append([role, dept, count])
    for col in "ABC":
        ws2.column_dimensions[col].width = 18

    path = "/app/utenti_ftchub.xlsx"
    wb.save(path)
    print(f"Salvato: {path} — {len(users)} utenti")

asyncio.run(main())
