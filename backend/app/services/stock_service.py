import csv
import io
import json
import re
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from openpyxl import Workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.stock import StockItem, StockSession, StockStoreData


_DATE_RE = re.compile(r"Stock-(\d{4}-\d{2}-\d{2})-IT0\d\.csv", re.IGNORECASE)

# Colonne extra per entity — non rappresentano negozi reali
_EXCLUDED_STORE_COLS: dict[str, set[str]] = {
    "IT01": {"IT105A"},
    "IT03": {"IT131"},
}


# ---------------------------------------------------------------------------
# Header-only parse — legge solo la prima riga del CSV
# ---------------------------------------------------------------------------

def parse_stock_header(file_content: bytes, filename: str, entity: str) -> tuple[date, list[str]]:
    """
    Legge solo l'header del CSV (prima riga).
    Ritorna (stock_date, store_codes).
    Nessun loop su righe dati — velocissimo.
    """
    m = _DATE_RE.search(filename)
    if not m:
        raise ValueError(
            f"Nome file non valido — formato atteso: Stock-YYYY-MM-DD-IT0X.csv (ricevuto: {filename})"
        )
    stock_date = date.fromisoformat(m.group(1))

    excluded = _EXCLUDED_STORE_COLS.get(entity, set())

    # Leggi solo la prima riga
    first_line = file_content.decode("latin-1").split("\n")[0]
    cols = [c.strip() for c in first_line.split(";")]
    store_codes = [c for c in cols[4:] if c and c not in excluded]

    return stock_date, store_codes


# ---------------------------------------------------------------------------
# Save — PostgreSQL COPY FROM STDIN + SQL puro, zero loop Python sui dati
# ---------------------------------------------------------------------------

async def save_stock_data(
    db: AsyncSession,
    file_content: bytes,
    entity: str,
    filename: str,
    stock_date: date,
    store_codes: list[str],
    user_id,
) -> StockSession:

    # Elimina sessione esistente (cascade su items + store_data)
    await db.execute(
        text("DELETE FROM ho.stock_sessions WHERE entity = :e AND stock_date = :d"),
        {"e": entity, "d": stock_date},
    )

    # Crea sessione
    session = StockSession(
        entity=entity,
        stock_date=stock_date,
        filename=filename,
        source="manual",
        total_items=0,  # aggiornato dopo COPY
        total_stores=len(store_codes),
        store_codes_json=json.dumps(store_codes),
        uploaded_by=user_id,
    )
    db.add(session)
    await db.flush()

    # Recupera la connessione asyncpg raw
    sa_conn = await db.connection()
    raw = await sa_conn.get_raw_connection()
    pg = raw.driver_connection

    # -----------------------------------------------------------------------
    # 1. Normalizza CSV: tronca ogni riga a esattamente n_expected colonne.
    #    Usa csv.reader per gestire correttamente quoting, trailing ';', ecc.
    # -----------------------------------------------------------------------
    excluded = _EXCLUDED_STORE_COLS.get(entity, set())
    in_reader = csv.reader(io.StringIO(file_content.decode("latin-1")), delimiter=";")
    out_buf = io.StringIO()
    out_writer = csv.writer(out_buf, delimiter=";", lineterminator="\r\n")

    # Indici delle colonne da tenere (calcolati sull'header del file raw)
    keep_indices: list[int] | None = None

    for i, row in enumerate(in_reader):
        if i == 0:
            # Costruisci la mappa degli indici da mantenere escludendo le colonne extra
            keep_indices = [
                j for j, col in enumerate(row)
                if col.strip() not in excluded
            ]
            out_writer.writerow([row[j] for j in keep_indices])
        else:
            if not any(f.strip() for f in row):
                continue  # salta righe vuote
            out_writer.writerow([row[j] if j < len(row) else "" for j in keep_indices])
    csv_content = out_buf.getvalue().encode("latin-1")

    # -----------------------------------------------------------------------
    # 2. Crea tabella staging con colonne TEXT (nessun problema di casting)
    # -----------------------------------------------------------------------
    store_col_defs = ", ".join(f'"{c}" TEXT' for c in store_codes)
    await pg.execute(f"""
        CREATE TEMP TABLE _stock_staging (
            item_no TEXT,
            description TEXT,
            description_local TEXT,
            adm TEXT,
            {store_col_defs}
        ) ON COMMIT DROP
    """)

    # -----------------------------------------------------------------------
    # 3. COPY CSV normalizzato direttamente in PostgreSQL
    # -----------------------------------------------------------------------
    await pg.copy_to_table(
        "_stock_staging",
        source=io.BytesIO(csv_content),
        format="csv",
        delimiter=";",
        header=True,
        encoding="latin1",
        null="",          # stringhe vuote → NULL (gestite con COALESCE)
    )

    # -----------------------------------------------------------------------
    # 3. INSERT stock_items via SQL puro
    # -----------------------------------------------------------------------
    result = await pg.fetch(f"""
        INSERT INTO ho.stock_items
            (session_id, item_no, description, description_local, adm_stock)
        SELECT
            {session.id},
            item_no,
            COALESCE(description, ''),
            COALESCE(description_local, ''),
            COALESCE(NULLIF(REPLACE(adm, ',', '.'), '')::NUMERIC::INTEGER, 0)
        FROM _stock_staging
        WHERE item_no IS NOT NULL AND item_no <> ''
        RETURNING id
    """)
    n_items = len(result)

    # -----------------------------------------------------------------------
    # 4. UNPIVOT store_data via SQL — una singola query, tutto in PostgreSQL
    # -----------------------------------------------------------------------
    if store_codes:
        values_clause = ", ".join(
            f"('{c}', COALESCE(NULLIF(REPLACE(ss.\"{c}\", ',', '.'), '')::NUMERIC::INTEGER, 0))"
            for c in store_codes
        )
        await pg.execute(f"""
            INSERT INTO ho.stock_store_data (item_id, store_code, quantity)
            SELECT si.id, v.store_code, v.quantity
            FROM ho.stock_items si
            JOIN _stock_staging ss
                ON si.item_no = ss.item_no
               AND si.session_id = {session.id}
            CROSS JOIN LATERAL (VALUES {values_clause}) AS v(store_code, quantity)
            WHERE v.quantity <> 0
        """)

    # Aggiorna contatore articoli nella sessione
    await pg.execute(
        "UPDATE ho.stock_sessions SET total_items = $1 WHERE id = $2",
        n_items, session.id,
    )

    await db.commit()
    await db.refresh(session)
    session.total_items = n_items
    return session


# ---------------------------------------------------------------------------
# Cleanup sessioni vecchie
# ---------------------------------------------------------------------------

async def cleanup_old_sessions(db: AsyncSession, entity: str, retention_days: int = 30) -> None:
    cutoff = datetime.now(timezone.utc) - timedelta(days=retention_days)
    await db.execute(
        text("""
            DELETE FROM ho.stock_sessions
            WHERE entity = :entity AND created_at < :cutoff
        """),
        {"entity": entity, "cutoff": cutoff},
    )
    await db.commit()


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

async def get_alldata_stats(db: AsyncSession) -> dict:
    result = await db.execute(text("""
        SELECT
            COUNT(*)                                              AS total_items,
            COUNT(prezzo_promo)                                   AS items_with_promo,
            COUNT(prezzo_bf)                                      AS items_with_bf,
            COUNT(eccezione_tipo)                                 AS items_with_eccezione,
            COALESCE(SUM(CASE WHEN is_bestseller THEN 1 ELSE 0 END), 0) AS items_bestseller,
            COALESCE(SUM(CASE WHEN is_scrap_inv  THEN 1 ELSE 0 END), 0) AS items_scrap_inv,
            COALESCE(SUM(CASE WHEN is_scrap_wd   THEN 1 ELSE 0 END), 0) AS items_scrap_wd,
            COALESCE(SUM(CASE WHEN is_picking     THEN 1 ELSE 0 END), 0) AS items_picking
        FROM ho.v_alldata
    """))
    row = result.mappings().one_or_none()

    stock_info: dict[str, dict] = {}
    for entity in ("IT01", "IT02", "IT03"):
        r = await db.execute(
            select(StockSession.stock_date, StockSession.store_codes_json)
            .where(StockSession.entity == entity)
            .order_by(StockSession.stock_date.desc())
            .limit(1)
        )
        sess = r.first()
        if sess:
            stock_info[entity] = {
                "date":        str(sess.stock_date),
                "store_codes": json.loads(sess.store_codes_json) if sess.store_codes_json else [],
            }
        else:
            stock_info[entity] = {"date": None, "store_codes": []}

    return {
        "total_items":          int(row["total_items"])          if row else 0,
        "items_with_promo":     int(row["items_with_promo"])     if row else 0,
        "items_with_bf":        int(row["items_with_bf"])        if row else 0,
        "items_with_eccezione": int(row["items_with_eccezione"]) if row else 0,
        "items_bestseller":     int(row["items_bestseller"])     if row else 0,
        "items_scrap_inv":      int(row["items_scrap_inv"])      if row else 0,
        "items_scrap_wd":       int(row["items_scrap_wd"])       if row else 0,
        "items_picking":        int(row["items_picking"])        if row else 0,
        "stock_date_it01":      stock_info["IT01"]["date"],
        "stock_date_it02":      stock_info["IT02"]["date"],
        "stock_date_it03":      stock_info["IT03"]["date"],
        "store_codes_it01":     stock_info["IT01"]["store_codes"],
        "store_codes_it02":     stock_info["IT02"]["store_codes"],
        "store_codes_it03":     stock_info["IT03"]["store_codes"],
    }


async def export_alldata_xlsx(db: AsyncSession) -> io.BytesIO:
    result = await db.execute(text("SELECT * FROM ho.v_alldata ORDER BY item_no"))
    rows = result.mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "ALLDATA"

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # Raccogli tutti i codici negozio presenti nei JSONB
    stores: dict[str, set[str]] = {"IT01": set(), "IT02": set(), "IT03": set()}
    for row in rows:
        for ent, key in (("IT01", "stock_it01"), ("IT02", "stock_it02"), ("IT03", "stock_it03")):
            if row[key]:
                stores[ent].update(row[key].keys())
    sc = {e: sorted(v) for e, v in stores.items()}

    headers = [
        "No.", "Description", "Description Local", "Warehouse",
        "Last Cost", "Unit Price", "Item Cat", "Net Weight", "Barcode",
        "VAT Code", "Units/Pack", "Model Store", "Batteries", "First RP",
        "Category", "Barcode Ext", "VAT %", "GM %",
        "Description1", "Description2", "Modulo", "Model Store Portale",
        "Modulo Numerico", "Model Store Portale Num",
        "Prezzo Promo", "Costo Promo", "Prezzo BF", "Costo BF",
        "Ecc. Prezzo 1", "Ecc. Prezzo 2", "Ecc. Sconto",
        "Ecc. Testo", "Ecc. Testo2", "Ecc. Categoria", "Ecc. Tipo",
        "BestSeller", "Scrap Inv", "Scrap WD", "Picking",
        "ADM IT01", "ADM IT02", "ADM IT03",
    ]
    headers += [f"IT01 {s}" for s in sc["IT01"]]
    headers += [f"IT02 {s}" for s in sc["IT02"]]
    headers += [f"IT03 {s}" for s in sc["IT03"]]
    ws.append(headers)

    def _n(v):
        return float(v) if v is not None else None

    for row in rows:
        fixed = [
            row["item_no"], row["description"], row["description_local"],
            row["warehouse"], _n(row["last_cost"]), _n(row["unit_price"]),
            row["item_cat"], _n(row["net_weight"]), row["barcode"],
            row["vat_code"], row["units_per_pack"], row["model_store"],
            row["batteries"], row["first_rp"], row["category"], row["barcode_ext"],
            _n(row["vat_pct"]), _n(row["gm_pct"]),
            row["description1"], row["description2"], row["modulo"],
            row["model_store_portale"], _n(row["modulo_numerico"]), _n(row["model_store_portale_num"]),
            _n(row["prezzo_promo"]), _n(row["costo_promo"]),
            _n(row["prezzo_bf"]),    _n(row["costo_bf"]),
            _n(row["eccezione_prezzo_1"]), _n(row["eccezione_prezzo_2"]),
            row["eccezione_sconto"], row["eccezione_testo"], row["eccezione_testo2"],
            row["eccezione_categoria"], row["eccezione_tipo"],
            row["is_bestseller"], row["is_scrap_inv"], row["is_scrap_wd"], row["is_picking"],
            row["adm_it01"], row["adm_it02"], row["adm_it03"],
        ]
        s1 = row["stock_it01"] or {}
        s2 = row["stock_it02"] or {}
        s3 = row["stock_it03"] or {}
        fixed += [s1.get(s, 0) for s in sc["IT01"]]
        fixed += [s2.get(s, 0) for s in sc["IT02"]]
        fixed += [s3.get(s, 0) for s in sc["IT03"]]
        ws.append(fixed)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_stock_xlsx(
    db: AsyncSession,
    session_id: int,
    store_code: Optional[str] = None,
) -> io.BytesIO:
    sess_result = await db.execute(
        select(StockSession).where(StockSession.id == session_id)
    )
    sess = sess_result.scalar_one_or_none()
    if not sess:
        raise ValueError("Sessione non trovata")

    all_stores = json.loads(sess.store_codes_json) if sess.store_codes_json else []

    if store_code:
        # Export filtrato per singolo store: NON includiamo la colonna ADM
        # (lo stock ADM non e' rilevante quando si guarda un solo negozio).
        rows = await db.execute(
            text("""
                SELECT si.item_no, si.description, si.description_local,
                       COALESCE(sd.quantity, 0) AS quantity
                FROM ho.stock_items si
                LEFT JOIN ho.stock_store_data sd
                    ON sd.item_id = si.id AND sd.store_code = :sc
                WHERE si.session_id = :sid
                ORDER BY si.item_no
            """),
            {"sid": session_id, "sc": store_code},
        )
        wb = Workbook()
        ws = wb.active
        ws.title = f"Stock {sess.entity} {sess.stock_date}"
        ws.append(["No.", "Description", "Description Local", store_code])
        for r in rows:
            ws.append(list(r))
    else:
        # Fetch all items + their store quantities in one query using JSON aggregation
        rows = await db.execute(
            text("""
                SELECT si.item_no, si.description, si.description_local, si.adm_stock,
                       json_object_agg(sd.store_code, sd.quantity)
                FROM ho.stock_items si
                LEFT JOIN ho.stock_store_data sd ON sd.item_id = si.id
                WHERE si.session_id = :sid
                GROUP BY si.id
                ORDER BY si.item_no
            """),
            {"sid": session_id},
        )
        wb = Workbook()
        ws = wb.active
        ws.title = f"Stock {sess.entity} {sess.stock_date}"
        ws.append(["No.", "Description", "Description Local", "ADM"] + all_stores)
        for item_no, desc, desc_local, adm, stores_json in rows:
            stores_map = stores_json or {}
            ws.append([item_no, desc, desc_local, adm] + [stores_map.get(s, 0) for s in all_stores])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

