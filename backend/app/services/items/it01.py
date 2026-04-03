import random
import string
from io import BytesIO
from typing import List

from openpyxl import load_workbook
from sqlalchemy import cast, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.types import String

from app.models.items import ItemImportSession, ItemMasterIT01

# Mappa nome colonna Excel → campo DB
COLUMN_MAP = {
    "Nr.":                          "item_no",
    "Descrizione":                  "description",
    "Description in Local Language":"description_local",
    "Magazzino":                    "warehouse",
    "Ultimo costo diretto":         "last_cost",
    "Prezzo unitario":              "unit_price",
    "Codice categoria articolo":    "item_cat",
    "Peso netto":                   "net_weight",
    "Barcode":                      "barcode",
    "Cat. reg. art./serv. IVA":     "vat_code",
    "Unità per collo":              "units_per_pack",
    "MODEL STORE":                  "model_store",
    "Batterie":                     "batteries",
    "First RP":                     "first_rp",
    "Category":                     "category",
    "Barcode ext.":                 "barcode_ext",
    "IVA":                          "vat_pct",
    "GM% escl. Trasporto":          "gm_pct",
    "Descrizione1":                 "description1",
    "Descrizione2":                 "description2",
    "Modulo":                       "modulo",
    "Model store per portale":      "model_store_portale",
    "MODULO NUMERICO":              "modulo_numerico",
    "MODEL STORE PORTALE NUMERICO": "model_store_portale_num",
}

BIGINT_FIELDS = {"barcode", "barcode_ext"}
INT_FIELDS = {"units_per_pack"}
DECIMAL_FIELDS = {"last_cost", "unit_price", "net_weight", "vat_pct", "gm_pct", "modulo_numerico", "model_store_portale_num"}
STRING_FIELDS = {
    "item_no", "description", "description_local", "warehouse",
    "item_cat", "vat_code", "model_store", "batteries",
    "first_rp", "category", "description1", "description2",
    "modulo", "model_store_portale",
}


def _random_batch_id() -> str:
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=8))


def _to_bigint(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(round(float(str(val))))
    except (ValueError, TypeError):
        return None


def _to_decimal(val) -> float | None:
    """
    Gestisce valori numerici restituiti da openpyxl.
    Le celle percentuale in Excel vengono già restituite come decimale (es. 0.22 per 22%).
    """
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except (ValueError, TypeError):
        return None


def _to_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_xlsx(file_bytes: bytes) -> List[dict]:
    """
    Legge il foglio 'ITEM LIST' del converter e restituisce una lista di dizionari
    pronti per l'inserimento in item_master_it01.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if "ITEM LIST" not in wb.sheetnames:
        raise ValueError('Foglio "ITEM LIST" non trovato nel file')

    ws = wb["ITEM LIST"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Foglio ITEM LIST vuoto o privo di dati")

    # Mappa nome colonna → indice posizione
    raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(raw_headers)}

    items = []
    for row in rows[1:]:
        # Salta righe vuote (prime 5 celle tutte None/"")
        if not row or all(row[j] in (None, "") for j in range(min(5, len(row)))):
            continue

        def get(col_name):
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        item_no = _to_str(get("Nr."))
        if not item_no:
            continue

        record = {}
        for xl_name, db_field in COLUMN_MAP.items():
            raw = get(xl_name)
            if db_field in BIGINT_FIELDS:
                record[db_field] = _to_bigint(raw)
            elif db_field in INT_FIELDS:
                record[db_field] = _to_int(raw)
            elif db_field in DECIMAL_FIELDS:
                record[db_field] = _to_decimal(raw)
            else:
                record[db_field] = _to_str(raw) or ""

        # description e description_local non possono essere None
        record["description"] = record.get("description") or ""
        record["description_local"] = record.get("description_local") or ""

        items.append(record)

    wb.close()
    return items


async def import_items_it01(
    file_bytes: bytes,
    source_filename: str,
    imported_by,
    db: AsyncSession,
) -> ItemImportSession:
    """
    Parsifica il file XLSX del converter, crea una nuova sessione IT01 (is_current=True)
    e inserisce tutti gli articoli in item_master_it01.
    La sessione precedente viene marcata is_current=False.
    """
    items_data = _parse_xlsx(file_bytes)
    if not items_data:
        raise ValueError("Nessun articolo trovato nel file")

    batch_id = _random_batch_id()

    # Demarca la sessione corrente precedente
    await db.execute(
        update(ItemImportSession)
        .where(ItemImportSession.entity == "IT01", ItemImportSession.is_current.is_(True))
        .values(is_current=False)
    )

    # Crea nuova sessione
    session = ItemImportSession(
        entity="IT01",
        imported_by=imported_by,
        batch_id=batch_id,
        source_filename=source_filename,
        is_current=True,
        row_count=0,
    )
    db.add(session)
    await db.flush()  # ottieni session.id

    # Inserimento bulk
    await db.execute(
        ItemMasterIT01.__table__.insert(),
        [{"session_id": session.id, **d} for d in items_data],
    )

    session.row_count = len(items_data)
    await db.commit()
    await db.refresh(session)
    return session


async def get_sessions_it01(db: AsyncSession) -> list[ItemImportSession]:
    result = await db.execute(
        select(ItemImportSession)
        .where(ItemImportSession.entity == "IT01")
        .order_by(ItemImportSession.imported_at.desc())
    )
    return result.scalars().all()


_ALLOWED_SORT = {
    "item_no", "description", "description_local", "unit_price",
    "barcode", "model_store", "category", "vat_pct",
    "units_per_pack", "description2",
}


async def get_items_it01(
    db: AsyncSession,
    session_id: int,
    search: str | None = None,
    model_store: str | None = None,
    category: str | None = None,
    sort_by: str = "item_no",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
) -> dict:
    if sort_by not in _ALLOWED_SORT:
        sort_by = "item_no"

    col = getattr(ItemMasterIT01, sort_by)
    order = col.asc() if sort_dir == "asc" else col.desc()

    base_q = select(ItemMasterIT01).where(ItemMasterIT01.session_id == session_id)

    if search and search.strip():
        like = f"%{search.strip()}%"
        base_q = base_q.where(
            or_(
                ItemMasterIT01.item_no.ilike(like),
                ItemMasterIT01.description.ilike(like),
                ItemMasterIT01.description_local.ilike(like),
                cast(ItemMasterIT01.barcode, String).ilike(like),
            )
        )

    if model_store and model_store.strip():
        base_q = base_q.where(ItemMasterIT01.model_store.ilike(f"%{model_store.strip()}%"))

    if category and category.strip():
        base_q = base_q.where(ItemMasterIT01.category.ilike(f"%{category.strip()}%"))

    total_result = await db.execute(
        select(func.count()).select_from(base_q.subquery())
    )
    total = total_result.scalar_one()

    offset = (page - 1) * page_size
    rows = (
        await db.execute(base_q.order_by(order).offset(offset).limit(page_size))
    ).scalars().all()

    return {"items": rows, "total": total}
