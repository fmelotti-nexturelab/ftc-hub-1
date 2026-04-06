"""
Genera i file Excel di archivio ItemList IT01 a partire dai dati in DB.

File generati:
 1. tbl_ItemM.xlsm  -> [storage]/02_ItemList/tbl_ItemM.xlsm  (sovrascritta ogni import)
 2. yyyymmdd - ItemM.xlsx -> [storage]/02_ItemList/Archivio/
 3. yyyymmdd_ItemListPortale.xlsx -> [storage]/02_ItemList/Archivio/
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.items import ItemImportSession, ItemMasterIT01
from app.services.app_settings_service import get_storage_path

logger = logging.getLogger(__name__)

# ── Definizione colonne ─────────────────────────────────────────────────────

FULL_HEADERS = [
    "Nr.", "Descrizione", "Description in Local Language", "Magazzino",
    "Ultimo costo diretto", "Prezzo unitario", "Codice categoria articolo",
    "Peso netto", "Barcode", "Cat. reg. art./serv. IVA", "Unit\u00e0 per collo",
    "MODEL STORE", "Batterie", "First RP", "Category", "Barcode ext.",
    "IVA", "GM% escl. Trasporto", "Descrizione1", "Descrizione2",
    None,  # colonna vuota (col U)
    "Modulo",
    "Model store per portale",
    "MODULO NUMERICO",
    "MODEL STORE PORTALE NUMERICO",
]

FULL_FIELD_MAP = [
    "item_no", "description", "description_local", "warehouse",
    "last_cost", "unit_price", "item_cat", "net_weight",
    "barcode", "vat_code", "units_per_pack", "model_store",
    "batteries", "first_rp", "category", "barcode_ext",
    "vat_pct", "gm_pct", "description1", "description2",
    None,  # colonna vuota
    "modulo",
    "model_store_portale",
    "modulo_numerico",
    "model_store_portale_num",
]


def _to_float(val):
    """Converti Decimal in float per openpyxl."""
    if val is not None and hasattr(val, "is_finite"):
        return float(val)
    return val


def _item_to_row(item, field_map):
    """Converte un record ItemMasterIT01 in una lista di valori per il file Excel."""
    row = []
    for field in field_map:
        if field is None:
            row.append(None)
        else:
            row.append(_to_float(getattr(item, field, None)))
    return row



# Formattazione colonne per ItemList IT01 archivio (0-based)
# D=3, G=6, K=10, Q=16, X=23, Y=24 → intero
# E=4, N=13 → decimale 2 cifre
# R=17 → float (nessuna forzatura decimali)
_IT01_INT_COLS = {0, 3, 6, 10, 16, 23, 24}
_IT01_DEC2_COLS = {4, 13}
_IT01_FLOAT_COLS = {17}


def _to_int_safe(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return val


def _write_xlsx(path: Path, sheet_name: str, headers: list, rows: list[list]):
    """Scrive un file .xlsx con formattazione colonne IT01."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        converted = []
        for col_idx, val in enumerate(row):
            if col_idx in _IT01_INT_COLS:
                converted.append(_to_int_safe(val))
            elif col_idx in _IT01_DEC2_COLS:
                converted.append(round(float(val), 2) if val is not None else None)
            elif col_idx in _IT01_FLOAT_COLS:
                converted.append(float(val) if val is not None else None)
            else:
                converted.append(val)
        ws.append(converted)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    logger.info("Scritto %s (%d righe)", path, len(rows))



TEMPLATES_DIR = Path(__file__).parent / "templates"

# Colonne da convertire in numero nel foglio ITEMS (0-based)
_TBL_NUM_COLS = [0, 3, 4, 5, 6, 7, 10, 13, 16, 17]  # A, D, E, F, G, H, K, N, Q, R
# Colonne con decimali (prezzo, peso, percentuali) — formato "0.00"
_TBL_DECIMAL_COLS = {4, 5, 7, 17}  # E(last_cost), F(unit_price), H(net_weight), R(gm_pct)


def _write_tbl_xlsm(path: Path, batch_id: str, headers: list, rows: list[list]):
    """
    Genera tbl_ItemM.xlsm con xlsxwriter (supporto nativo .xlsm).

    - Foglio "ITEMS": B1=batch_id, riga 2=header, dalla riga 3=dati (col A-V)
    - Foglio "data": B1=data oggi, C1=ora (hh:mm), B2="week XX", B5=batch_id
    """
    import xlsxwriter
    from zoneinfo import ZoneInfo

    now = datetime.now(ZoneInfo("Europe/Rome"))
    max_col = 22  # Colonne A-V

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = xlsxwriter.Workbook(str(path))

    # ── Foglio "ITEMS" ────────────────────────────────────────────────────────
    ws = wb.add_worksheet("ITEMS")

    # Riga 0 (Excel riga 1): B1 = batch_id
    ws.write(0, 1, batch_id)

    # Riga 1 (Excel riga 2): header (colonne A-V)
    for col_idx, val in enumerate(headers[:max_col]):
        ws.write(1, col_idx, val)

    # Dalla riga 2 (Excel riga 3): dati (colonne A-V)
    for row_idx, row in enumerate(rows):
        for col_idx in range(min(max_col, len(row))):
            val = row[col_idx]
            excel_row = row_idx + 2  # offset: riga 0=batch, riga 1=header
            if col_idx in _TBL_NUM_COLS and val is not None:
                try:
                    ws.write_number(excel_row, col_idx, float(str(val)))
                except (ValueError, TypeError):
                    ws.write(excel_row, col_idx, val)
            else:
                ws.write(excel_row, col_idx, val)

    # ── Foglio "data" ─────────────────────────────────────────────────────────
    ws_data = wb.add_worksheet("data")
    date_fmt = wb.add_format({"num_format": "dd-mm-yyyy"})
    ws_data.write_string(0, 0, "data ultima modifica")                # A1
    ws_data.write_string(1, 0, "messaggio")                           # A2
    ws_data.write_string(4, 0, "matricola")                           # A5
    ws_data.write_datetime(0, 1, now.replace(tzinfo=None), date_fmt)  # B1 = data
    ws_data.write_string(0, 2, now.strftime("%H:%M"))                 # C1 = ora
    ws_data.write_string(1, 1, f"week {now.isocalendar()[1]}")        # B2 = week
    ws_data.write_string(4, 1, batch_id)                              # B5 = batch_id

    wb.close()
    logger.info("Scritto %s (%d righe)", path, len(rows))


# Indici colonne nel full row (0-based, corrispondono a A=0, B=1, ... Y=24)
_COL_U = 20  # colonna vuota
_COL_X = 23  # MODULO NUMERICO
_COL_Y = 24  # MODEL STORE PORTALE NUMERICO

# Indici colonne nel portale (sheet 1, 12 colonne A-L, 0-based)
_P_A = 0   # Nr.
_P_B = 1   # Descrizione
_P_C = 2   # Description in Local Language
_P_D = 3   # Magazzino
_P_H = 7   # Peso netto
_P_I = 8   # Barcode
_P_G = 6   # Codice categoria articolo
_P_K = 10  # Unita per collo


PORTALE_EXTRA_ROWS = [
    ["armadietto",  "armadietto",  "armadietto",  1, 1, 1, 1, 0, "armadietto",  "BTW22", 0, 0],
    ["table",       "table",       "table",       1, 1, 1, 1, 0, "table",       "BTW22", 0, 0],
    ["dogs",        "dogs",        "dogs",        1, 1, 1, 1, 0, "dogs",        "BTW22", 0, 0],
    ["extra",       "extra",       "extra",       1, 1, 1, 1, 0, "extra",       "BTW22", 0, 0],
    ["magazino",    "magazino",    "magazino",    1, 1, 1, 1, 0, "magazino1",   "BTW22", 0, 0],
    ["pallet",      "pallet",      "pallet",      1, 1, 1, 1, 0, "pallet",      "BTW22", 0, 0],
    ["reklamation", "reklamation", "reklamation", 1, 1, 1, 1, 0, "reklamation", "BTW22", 0, 0],
    ["slatwall",    "slatwall",    "slatwall",    1, 1, 1, 1, 0, "slatwall",    "BTW22", 0, 0],
]


def _write_portale_xlsx(path: Path, full_rows: list[list]):
    """
    Genera il file ItemListPortale per il portale fatture a partire dal template.

    Processo:
    1. Apre il template, svuota sheet 1 da riga 2
    2. Crea sheet 2 con tutti i dati ItemM (25 colonne, solo dati)
    3. Popola sheet 1 (12 colonne) dai primi 12 campi, poi sovrascrive:
       K ← col U,  H ← col Y,  D ← col X
    4. Imposta K1 = "Unita per collo"
    5. Rimuove righe con Barcode (col I) vuoto
    6. Converte Barcode in intero
    7. Riempie B↔C dove vuoti
    8. Col A in numero, D = 1, H = 1
    """
    template_path = TEMPLATES_DIR / "Itemlist_portale_template.xlsx"
    if not template_path.exists():
        raise ValueError("Template portale non trovato: " + str(template_path))

    wb = load_workbook(str(template_path))
    ws1 = wb.worksheets[0]

    # ── Step 1: svuota sheet 1 da riga 2 in giù ──────────────────────────────
    if ws1.max_row > 1:
        ws1.delete_rows(2, ws1.max_row - 1)

    # ── Step 2: crea sheet 2 con tutti i dati ItemM (solo righe dati) ─────────
    ws2 = wb.create_sheet("ItemM")
    for row in full_rows:
        ws2.append(row)

    # ── Step 3-4: costruisci righe portale con override colonne ────────────────
    portale_rows = []
    for fr in full_rows:
        pr = list(fr[:12])                       # prime 12 colonne (A-L)
        pr[_P_K] = fr[_COL_U] if len(fr) > _COL_U else None   # U → K
        pr[_P_H] = fr[_COL_Y] if len(fr) > _COL_Y else None   # Y → H
        pr[_P_D] = fr[_COL_X] if len(fr) > _COL_X else None   # X → D
        portale_rows.append(pr)

    # ── Step 5: rimuovi righe con Barcode vuoto ───────────────────────────────
    portale_rows = [
        r for r in portale_rows
        if r[_P_I] is not None and r[_P_I] != "" and r[_P_I] != 0
    ]

    # ── Step 6: converti Barcode (col I) in intero ────────────────────────────
    for row in portale_rows:
        if row[_P_I] is not None:
            try:
                row[_P_I] = int(float(str(row[_P_I])))
            except (ValueError, TypeError):
                pass

    # ── Step 7: riempi B↔C dove vuoti ─────────────────────────────────────────
    for row in portale_rows:
        b = str(row[_P_B]).strip() if row[_P_B] else ""
        c = str(row[_P_C]).strip() if row[_P_C] else ""
        if not b and c:
            row[_P_B] = row[_P_C]
        if not c and b:
            row[_P_C] = row[_P_B]

    # ── Step 8: col A e G in numero, D = 1, H = 1, K = 1 ──────────────────────
    for row in portale_rows:
        for ci in (_P_A, _P_G):
            if row[ci] is not None:
                try:
                    row[ci] = int(float(str(row[ci])))
                except (ValueError, TypeError):
                    pass
        row[_P_D] = 1
        row[_P_H] = 1
        row[_P_K] = 1

    # ── Step 4b: imposta header K1 ────────────────────────────────────────────
    ws1.cell(row=1, column=_P_K + 1, value="Unita per collo")

    # ── Scrivi righe in sheet 1 ───────────────────────────────────────────────
    for row in portale_rows:
        ws1.append(row)

    # ── Righe fisse in fondo (zone inventario) ────────────────────────────────
    for row in PORTALE_EXTRA_ROWS:
        ws1.append(row)

    # ── Formato numero intero (0 decimali) sulla colonna Barcode (col I) ──────
    for row_idx in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=row_idx, column=_P_I + 1)
        cell.number_format = "0"

    # ── Elimina sheet 2 (ItemM) — serviva solo come appoggio ──────────────────
    wb.remove(ws2)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    logger.info("Scritto portale fatture %s (%d righe)", path, len(portale_rows))
    return len(portale_rows)


async def generate_itemlist_files(
    session_id: int,
    db: AsyncSession,
) -> dict:
    """
    Genera i file Excel per una sessione ItemList IT01 e li salva
    nella cartella FTC HUB Storage configurata.
    """
    storage_path = await get_storage_path(db)
    base = Path(storage_path)

    if not base.exists():
        raise ValueError(f"La cartella FTC HUB Storage non esiste: {storage_path}")

    sess_result = await db.execute(
        select(ItemImportSession).where(ItemImportSession.id == session_id)
    )
    session = sess_result.scalar_one_or_none()
    if not session:
        raise ValueError(f"Sessione {session_id} non trovata")

    items_result = await db.execute(
        select(ItemMasterIT01)
        .where(ItemMasterIT01.session_id == session_id)
        .order_by(ItemMasterIT01.item_no)
    )
    items = items_result.scalars().all()
    if not items:
        raise ValueError("Nessun articolo nella sessione")

    full_rows = [_item_to_row(item, FULL_FIELD_MAP) for item in items]
    date_str = datetime.now().strftime("%Y%m%d")

    # 1. tbl_ItemM.xlsm (da template)
    tbl_path = base / "02_ItemList" / "tbl_ItemM.xlsm"
    _write_tbl_xlsm(tbl_path, session.batch_id, FULL_HEADERS, full_rows)

    # 2. yyyymmdd_IT01_ItemList.xlsx (archivio)
    itemm_name = f"{date_str}_IT01_ItemList.xlsx"
    itemm_path = base / "02_ItemList" / "Archivio" / itemm_name
    _write_xlsx(itemm_path, "Foglio1", FULL_HEADERS, full_rows)

    # 3. yyyymmdd_ItemListPortale.xlsx
    portale_name = f"{date_str}_ItemListPortale.xlsx"
    portale_path = base / "02_ItemList" / "Archivio" / portale_name
    portale_count = _write_portale_xlsx(portale_path, full_rows)

    logger.info(
        "Generati file ItemList: %d articoli totali, %d nel portale",
        len(items), portale_count,
    )

    return {
        "tbl_item_m": str(tbl_path),
        "item_m": str(itemm_path),
        "itemlist_portale": str(portale_path),
        "row_count": len(items),
        "portale_count": portale_count,
    }
