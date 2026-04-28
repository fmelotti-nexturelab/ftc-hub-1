from datetime import datetime, timezone, date
from typing import List, Optional, Set
import random

import logging

import io
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, cast, String

logger = logging.getLogger(__name__)

from app.database import get_db
from app.models.operator_code import OperatorCode, OperatorCodeRequest, OperatorCodePool
from app.models.stores import Store
from app.models.auth import User
from app.schemas.operator_code import (
    OperatorCodeResponse,
    OperatorCodeCreate,
    OperatorCodeRequestPayload,
    OperatorCodeRequestResponse,
    PendingRequestsResponse,
    BulkRequestPayload,
    BulkRowResult,
    EvadiPayload,
    BulkEvadiRow,
    BulkEvadiPayload,
    BulkEvadiResult,
    NotifyResultItem,
    NotifyResponse,
    NotifyPayload,
    PoolPreviewRow,
    PoolPreviewResponse,
    PoolImportPayload,
    BulkDeletePayload,
)
from app.schemas.tickets import TicketCreate
from app.models.tickets import TicketPriority, Ticket, TicketStatus
from app.core.dependencies import require_permission, get_current_user
from app.services.tickets import ticket_service
from app.services.nav_export import generate_nav_files

router = APIRouter(prefix="/api/ho/operator-codes", tags=["HO - Codice Operatore"])

_PERM_VIEW = require_permission("codici_operatore")

DOMAIN = "@flyingtigeritalia.com"
DIGITS_NO_ZERO = "123456789"


def _entity_from_store(store_number: str) -> Optional[str]:
    """Ricava l'entity dal prefisso del negozio (es. IT01055 -> IT01)."""
    s = (store_number or "").upper().strip()
    for e in ("IT01", "IT02", "IT03"):
        if s.startswith(e):
            return e
    return None


def _generate_password() -> str:
    """4 cifre casuali, nessuno zero."""
    return "".join(random.choices(DIGITS_NO_ZERO, k=4))


def _suggest_email(first_name: str, last_name: str, existing_emails: Set[str]) -> Optional[str]:
    fn = first_name.lower().replace(" ", "")
    ln = last_name.lower().replace(" ", "")

    def _try(prefix: str) -> Optional[str]:
        if len(prefix) < 2:
            return None
        full = prefix + DOMAIN
        return prefix if full not in existing_emails else None

    # Combinazioni in ordine: 3+3, 4+2, 5+1 — nome prima, poi cognome prima
    candidates = [
        fn[:3] + ln[:3],
        ln[:3] + fn[:3],
        fn[:4] + ln[:2],
        ln[:4] + fn[:2],
        fn[:5] + ln[:1],
        ln[:5] + fn[:1],
    ]
    for candidate in candidates:
        result = _try(candidate)
        if result is not None:
            return result
    return None


def _serialize_code(oc: OperatorCode, requester_name: Optional[str] = None, creator_name: Optional[str] = None) -> dict:
    return {
        "id": str(oc.id),
        "code": oc.code,
        "first_name": oc.first_name,
        "last_name": oc.last_name,
        "email": oc.email,
        "start_date": oc.start_date.isoformat() if oc.start_date else None,
        "store_number": oc.store_number,
        "created_at": oc.created_at.isoformat() if oc.created_at else None,
        "requested_at": oc.requested_at.isoformat() if oc.requested_at else None,
        "requester_name": requester_name,
        "creator_name": creator_name,
    }


async def _get_cumulative_ticket(db: AsyncSession) -> Optional[Ticket]:
    result = await db.execute(
        select(Ticket).where(
            Ticket.title == "Richiesta Codice Operatore",
            cast(Ticket.status, String) != TicketStatus.CLOSED.value,
            Ticket.is_active == True,
        ).order_by(Ticket.created_at.desc()).limit(1)
    )
    return result.scalar_one_or_none()




@router.get(
    "",
    response_model=List[OperatorCodeResponse],
    dependencies=[Depends(_PERM_VIEW)],
)
async def list_operator_codes(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(OperatorCode)
        .where(OperatorCode.is_active == True)
        .order_by(OperatorCode.last_name, OperatorCode.first_name)
    )
    return result.scalars().all()


@router.post(
    "/overwrite",
    dependencies=[Depends(require_permission("codici_operatore"))],
)
async def overwrite_operator_codes(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Sovrascrive ho.operator_codes con il contenuto di un file Excel."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File non valido: carica un file .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Impossibile leggere il file Excel")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    required = {"codice", "nome", "cognome"}
    missing = required - col.keys()
    if missing:
        raise HTTPException(400, f"Colonne mancanti nel file: {', '.join(missing)}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codice    = str(row[col["codice"]]).strip() if row[col["codice"]] is not None else None
        nome      = str(row[col["nome"]]).strip()   if row[col["nome"]] is not None else None
        cognome   = str(row[col["cognome"]]).strip() if row[col["cognome"]] is not None else None
        email     = str(row[col["email"]]).strip()   if "email" in col and row[col["email"]] else None
        store     = str(row[col["store"]]).strip()   if "store" in col and row[col["store"]] else None
        if not nome and not cognome:
            continue
        rows.append(OperatorCode(
            code=codice,
            first_name=nome,
            last_name=cognome,
            email=email,
            store_number=store,
            is_active=True,
        ))

    await db.execute(delete(OperatorCode))
    db.add_all(rows)
    await db.commit()
    return {"inserted": len(rows)}


@router.get(
    "/badge-count",
    dependencies=[Depends(_PERM_VIEW)],
)
async def badge_count(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(func.count()).select_from(OperatorCodeRequest)
        .where(OperatorCodeRequest.is_evaded == False)
    )
    return {"count": result.scalar() or 0}


@router.get(
    "/requests",
    dependencies=[Depends(_PERM_VIEW)],
)
async def list_requests(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    reqs_result = await db.execute(
        select(OperatorCodeRequest).order_by(
            OperatorCodeRequest.store_number,
            OperatorCodeRequest.last_name,
            OperatorCodeRequest.first_name,
        )
    )
    requests = reqs_result.scalars().all()

    # Carica tutte le email esistenti per il suggest
    emails_result = await db.execute(
        select(OperatorCode.email).where(
            OperatorCode.email.isnot(None),
            OperatorCode.is_active == True,
        )
    )
    existing_emails: Set[str] = {row[0].lower() for row in emails_result.all() if row[0]}

    items = []
    for req in requests:
        requester_name = None
        if req.requested_by:
            u = await db.execute(select(User).where(User.id == req.requested_by))
            user_obj = u.scalar_one_or_none()
            if user_obj:
                requester_name = user_obj.full_name
        suggested = _suggest_email(req.first_name, req.last_name, existing_emails)
        if suggested:
            # riserva l'email per le righe successive nella stessa batch
            existing_emails.add((suggested + DOMAIN).lower())
        items.append(OperatorCodeRequestResponse(
            id=req.id,
            first_name=req.first_name,
            last_name=req.last_name,
            store_number=req.store_number,
            start_date=req.start_date,
            notes=req.notes,
            requested_by=req.requested_by,
            requester_name=requester_name,
            created_at=req.created_at,
            is_evaded=req.is_evaded,
            evaded_at=req.evaded_at,
            suggested_email=suggested,
            assigned_code=req.assigned_code,
            assigned_password=req.assigned_password,
            assigned_email=req.assigned_email,
            exported_at=req.exported_at,
        ))

    ticket = await _get_cumulative_ticket(db)
    ticket_status = ticket.status.value if ticket else None

    return PendingRequestsResponse(items=items, ticket_status=ticket_status)


@router.post(
    "/requests/take-over",
    dependencies=[Depends(_PERM_VIEW)],
)
async def take_over_requests(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    ticket = await _get_cumulative_ticket(db)
    if not ticket:
        raise HTTPException(404, "Nessuna richiesta in attesa")

    if ticket.status == TicketStatus.OPEN:
        ticket.status = TicketStatus.IN_PROGRESS
        ticket.assigned_to = current_user.id
        ticket.taken_at = datetime.now(timezone.utc)
        await db.commit()

    return {"message": "Preso in carico"}


@router.post(
    "/requests/close-ticket",
    dependencies=[Depends(_PERM_VIEW)],
)
async def close_ticket(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    # 1. Carica tutte le richieste evase
    evaded_result = await db.execute(
        select(OperatorCodeRequest).where(OperatorCodeRequest.is_evaded == True)
    )
    evaded_reqs = evaded_result.scalars().all()
    total = len(evaded_reqs)

    # 2. Upsert in operator_codes
    inserted = 0
    for req in evaded_reqs:
        existing = await db.execute(
            select(OperatorCode).where(
                func.lower(OperatorCode.first_name) == req.first_name.lower(),
                func.lower(OperatorCode.last_name) == req.last_name.lower(),
                OperatorCode.is_active == True,
            )
        )
        op = existing.scalar_one_or_none()
        if op:
            if req.assigned_code:
                op.code = str(req.assigned_code)
            if req.assigned_email:
                op.email = req.assigned_email
            if req.store_number:
                op.store_number = req.store_number
            if req.start_date:
                op.start_date = req.start_date
        else:
            db.add(OperatorCode(
                first_name=req.first_name,
                last_name=req.last_name,
                store_number=req.store_number,
                start_date=req.start_date,
                email=req.assigned_email,
                code=str(req.assigned_code) if req.assigned_code else None,
                is_active=True,
                requested_by=req.requested_by,
            ))
            inserted += 1

    # 3. Svuota tutte le richieste
    await db.execute(delete(OperatorCodeRequest))
    await db.commit()

    # 4. Chiudi il ticket
    ticket = await _get_cumulative_ticket(db)
    if ticket:
        from app.schemas.tickets import TicketStatusUpdate
        await ticket_service.update_status(
            db, ticket.id, TicketStatusUpdate(status=TicketStatus.CLOSED), current_user
        )

    return {"message": "Ticket chiuso", "total": total, "inserted": inserted, "updated": total - inserted}


@router.post(
    "/requests/mark-notified",
    dependencies=[Depends(_PERM_VIEW)],
)
async def mark_requests_notified(
    body: BulkDeletePayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    result = await db.execute(
        select(OperatorCodeRequest).where(
            cast(OperatorCodeRequest.id, String).in_(body.ids)
        )
    )
    reqs = result.scalars().all()
    now = datetime.now(timezone.utc)
    for req in reqs:
        req.notification_sent_at = now
    await db.commit()
    return {"marked": len(reqs)}


@router.delete(
    "/requests/bulk-delete",
    dependencies=[Depends(_PERM_VIEW)],
)
async def bulk_delete_requests(
    body: BulkDeletePayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    result = await db.execute(
        select(OperatorCodeRequest).where(
            cast(OperatorCodeRequest.id, String).in_(body.ids)
        )
    )
    requests_to_delete = result.scalars().all()
    for req in requests_to_delete:
        await db.delete(req)

    await db.commit()

    return {"deleted": len(requests_to_delete)}


@router.delete(
    "/requests/{request_id}",
    dependencies=[Depends(_PERM_VIEW)],
)
async def process_request(
    request_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    result = await db.execute(
        select(OperatorCodeRequest).where(OperatorCodeRequest.id == request_id)
    )
    req = result.scalar_one_or_none()
    if not req:
        raise HTTPException(404, "Richiesta non trovata")

    await db.delete(req)
    await db.commit()



    return {"message": "Richiesta eliminata"}


@router.patch(
    "/requests/{request_id}/evadi",
    dependencies=[Depends(_PERM_VIEW)],
)
async def evadi_request(
    request_id: str,
    body: EvadiPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    result = await db.execute(
        select(OperatorCodeRequest).where(OperatorCodeRequest.id == request_id)
    )
    req = result.scalar_one_or_none()
    if not req:
        raise HTTPException(404, "Richiesta non trovata")

    # Determina entity dal negozio
    entity = _entity_from_store(req.store_number)
    if not entity:
        raise HTTPException(400, f"Impossibile determinare l'entity dal negozio '{req.store_number}'")

    # Calcola prossimo codice (max 4 cifre + 1)
    max_result = await db.execute(
        select(func.max(OperatorCodePool.code)).where(
            OperatorCodePool.entity == entity,
            OperatorCodePool.code >= 1000,
            OperatorCodePool.code <= 9999,
        )
    )
    max_code = max_result.scalar() or 1000
    next_code = max_code + 1

    # Genera password
    password = _generate_password()

    # Gestione email
    email_prefix = (body.email or "").strip()
    full_email = (email_prefix + DOMAIN) if email_prefix and not email_prefix.endswith(DOMAIN) else (email_prefix or None)

    # Aggiorna operator_codes se esiste
    if full_email:
        op_result = await db.execute(
            select(OperatorCode).where(
                func.lower(OperatorCode.first_name) == req.first_name.lower(),
                func.lower(OperatorCode.last_name) == req.last_name.lower(),
                OperatorCode.is_active == True,
            )
        )
        op = op_result.scalar_one_or_none()
        if op:
            op.email = full_email
            op.code = str(next_code)

    # Aggiungi codice al pool
    db.add(OperatorCodePool(entity=entity, code=next_code))

    # Marca richiesta come evasa con i dati assegnati
    req.is_evaded = True
    req.evaded_at = datetime.now(timezone.utc)
    req.assigned_code = next_code
    req.assigned_password = password
    req.assigned_email = full_email

    await db.commit()


    return {"message": "Richiesta evasa", "assigned_code": next_code, "assigned_password": password}


@router.post(
    "/requests/bulk-evadi",
    dependencies=[Depends(_PERM_VIEW)],
)
async def bulk_evadi_requests(
    body: BulkEvadiPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    results: list[dict] = []

    for row in body.rows:
        req_res = await db.execute(
            select(OperatorCodeRequest).where(OperatorCodeRequest.id == row.id)
        )
        req = req_res.scalar_one_or_none()

        if not req:
            results.append(BulkEvadiResult(
                id=row.id, first_name="?", last_name="?", store_number="?",
                status="error", note="Richiesta non trovata",
            ).model_dump())
            continue

        if req.is_evaded:
            results.append(BulkEvadiResult(
                id=row.id, first_name=req.first_name, last_name=req.last_name,
                store_number=req.store_number, status="error", note="Già evasa",
            ).model_dump())
            continue

        entity = _entity_from_store(req.store_number)
        if not entity:
            results.append(BulkEvadiResult(
                id=row.id, first_name=req.first_name, last_name=req.last_name,
                store_number=req.store_number, status="error",
                note=f"Entity non determinabile da '{req.store_number}'",
            ).model_dump())
            continue

        max_result = await db.execute(
            select(func.max(OperatorCodePool.code)).where(
                OperatorCodePool.entity == entity,
                OperatorCodePool.code >= 1000,
                OperatorCodePool.code <= 9999,
            )
        )
        max_code = max_result.scalar() or 1000
        next_code = max_code + 1
        password = _generate_password()

        email_prefix = (row.email or "").strip()
        full_email = (email_prefix + DOMAIN) if email_prefix and not email_prefix.endswith(DOMAIN) else (email_prefix or None)

        if full_email:
            op_result = await db.execute(
                select(OperatorCode).where(
                    func.lower(OperatorCode.first_name) == req.first_name.lower(),
                    func.lower(OperatorCode.last_name) == req.last_name.lower(),
                    OperatorCode.is_active == True,
                )
            )
            op = op_result.scalar_one_or_none()
            if op:
                op.email = full_email
                op.code = str(next_code)

        db.add(OperatorCodePool(entity=entity, code=next_code))
        req.is_evaded = True
        req.evaded_at = datetime.now(timezone.utc)
        req.assigned_code = next_code
        req.assigned_password = password
        req.assigned_email = full_email

        results.append(BulkEvadiResult(
            id=row.id, first_name=req.first_name, last_name=req.last_name,
            store_number=req.store_number, assigned_code=next_code,
            assigned_email=full_email, assigned_password=password, status="ok",
        ).model_dump())

    await db.commit()


    ok = sum(1 for r in results if r["status"] == "ok")
    return {"results": results, "ok": ok, "errors": len(results) - ok}


@router.post(
    "/notify",
    response_model=NotifyResponse,
    dependencies=[Depends(_PERM_VIEW)],
)
async def notify_operators(
    body: NotifyPayload = NotifyPayload(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    preview: bool = False,
    ids: Optional[str] = None,  # lista UUID separati da virgola
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    from app.services.tickets.notification_service import (
        notify_operator_code_assigned,
        _build_operator_code_email,
    )

    filters = [
        OperatorCodeRequest.is_evaded == True,
        OperatorCodeRequest.notification_sent_at.is_(None),
        OperatorCodeRequest.assigned_code.isnot(None),
    ]
    if ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        filters.append(cast(OperatorCodeRequest.id, String).in_(id_list))

    result = await db.execute(
        select(OperatorCodeRequest).where(*filters)
        .order_by(OperatorCodeRequest.store_number, OperatorCodeRequest.last_name)
    )
    to_notify = result.scalars().all()

    if not to_notify:
        return NotifyResponse(sent=0, skipped=0, results=[])

    sender_name = current_user.full_name or current_user.username
    sender_email = current_user.email
    override_map = {o.request_id: o for o in body.overrides}
    results: list[NotifyResultItem] = []
    sent_count = 0

    for req in to_notify:
        store_res = await db.execute(
            select(Store).where(
                func.lower(Store.store_number) == req.store_number.lower(),
                Store.is_active == True,
            )
        )
        store = store_res.scalar_one_or_none()

        # Applica override destinatari se presenti
        ov = override_map.get(str(req.id))
        sm_mail = (ov.sm_mail if ov and ov.sm_mail is not None else (store.sm_mail if store else None)) or None
        dm_mail = (ov.dm_mail if ov and ov.dm_mail is not None else (store.dm_mail if store else None)) or None
        sm_name = store.sm_name if store else None
        dm_name = store.dm_name if store else None

        item = NotifyResultItem(
            request_id=str(req.id),
            first_name=req.first_name,
            last_name=req.last_name,
            store_number=req.store_number,
            sm_name=sm_name,
            sm_mail=sm_mail,
            dm_name=dm_name,
            dm_mail=dm_mail,
        )

        if not store:
            item.error = "Negozio non trovato"
            results.append(item)
            continue

        if preview:
            # Genera solo HTML, non spedisce e non aggiorna il DB
            recipient_name = sm_name or dm_name or "Store Manager"
            item.html_preview = _build_operator_code_email(
                recipient_name=recipient_name,
                store_number=req.store_number,
                store_name=store.store_name or req.store_number,
                first_name=req.first_name,
                last_name=req.last_name,
                assigned_code=req.assigned_code,
                assigned_password=req.assigned_password or "",
                assigned_email=req.assigned_email,
                start_date=req.start_date,
            )
            item.sm_sent = bool(sm_mail)
            item.dm_sent = bool(dm_mail)
            results.append(item)
            continue

        try:
            await notify_operator_code_assigned(
                store_number=req.store_number,
                store_name=store.store_name or req.store_number,
                sm_name=sm_name,
                sm_mail=sm_mail,
                dm_name=dm_name,
                dm_mail=dm_mail,
                first_name=req.first_name,
                last_name=req.last_name,
                assigned_code=req.assigned_code,
                assigned_password=req.assigned_password or "",
                assigned_email=req.assigned_email,
                start_date=req.start_date,
                sender_name=sender_name,
                sender_email=sender_email,
            )
            item.sm_sent = bool(sm_mail)
            item.dm_sent = bool(dm_mail)
            req.notification_sent_at = datetime.now(timezone.utc)
            sent_count += 1
        except Exception as exc:
            item.error = str(exc)
            logger.warning(f"Notifica fallita per {req.store_number} / {req.last_name}: {exc}")

        results.append(item)

    await db.commit()
    skipped = len(to_notify) - sent_count
    return NotifyResponse(sent=sent_count, skipped=skipped, results=results)


@router.post(
    "/bulk-request",
    dependencies=[Depends(_PERM_VIEW)],
)
async def bulk_request(
    data: BulkRequestPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_type
    results: list[BulkRowResult] = []
    any_inserted = False

    for row in data.rows:
        fn = (row.first_name or "").strip()
        ln = (row.last_name or "").strip()
        sn = (row.store_number or "").strip().upper()
        sd_str = (row.start_date or "").strip()

        # Valida campi obbligatori
        if not fn or not ln or not sn or not sd_str:
            results.append(BulkRowResult(
                first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
                status="error", note="Dati incompleti",
            ))
            continue

        # Valida e converti data
        try:
            sd = date_type.fromisoformat(sd_str)
        except ValueError:
            results.append(BulkRowResult(
                first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
                status="error", note="Data non valida",
            ))
            continue

        # Valida negozio
        store_res = await db.execute(
            select(Store).where(func.lower(Store.store_number) == sn.lower(), Store.is_active == True)
        )
        if not store_res.scalar_one_or_none():
            results.append(BulkRowResult(
                first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
                status="error", note=f"Negozio '{sn}' non trovato",
            ))
            continue

        # Controlla operator_codes
        exact_res = await db.execute(
            select(OperatorCode).where(
                func.lower(OperatorCode.first_name) == fn.lower(),
                func.lower(OperatorCode.last_name) == ln.lower(),
                OperatorCode.is_active == True,
            )
        )
        exact = exact_res.scalar_one_or_none()
        if exact:
            note = f"Già presente nel sistema"
            if exact.code:
                note += f" (codice: {exact.code})"
            results.append(BulkRowResult(
                first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
                status="exists", note=note,
            ))
            continue

        # Controlla operator_code_requests
        pending_res = await db.execute(
            select(OperatorCodeRequest).where(
                func.lower(OperatorCodeRequest.first_name) == fn.lower(),
                func.lower(OperatorCodeRequest.last_name) == ln.lower(),
            )
        )
        if pending_res.scalar_one_or_none():
            results.append(BulkRowResult(
                first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
                status="pending", note="Richiesta già in attesa di gestione",
            ))
            continue

        # Inserisci
        db.add(OperatorCodeRequest(
            first_name=fn, last_name=ln, store_number=sn,
            start_date=sd, requested_by=current_user.id,
        ))
        any_inserted = True
        results.append(BulkRowResult(
            first_name=fn, last_name=ln, store_number=sn, start_date=sd_str,
            status="inserted", note="Richiesta inserita",
        ))

    if any_inserted:
        await db.commit()
        existing_ticket = await _get_cumulative_ticket(db)
        if not existing_ticket:
            requester_name = current_user.full_name or current_user.username
            inserted_count = sum(1 for r in results if r.status == "inserted")
            try:
                inserted_names = "\n".join(
                    f"  - {r.last_name} {r.first_name} — {r.store_number}"
                    for r in results if r.status == "inserted"
                )
                ticket_data = TicketCreate(
                    title="Richiesta Codice Operatore",
                    description=(
                        f"Import massivo: {inserted_count} nuove richieste codice operatore in attesa di gestione.\n"
                        f"Inoltrate da: {requester_name}\n\n"
                        f"Operatori:\n{inserted_names}"
                    ),
                    category_id=7,
                    subcategory_id=17,
                    priority=TicketPriority.MEDIUM,
                    requester_name=requester_name,
                    requester_email=current_user.email or "",
                    requester_phone="",
                    teamviewer_code="",
                )
                await ticket_service.create_ticket(db, ticket_data, current_user)
            except Exception:
                pass

    return {"results": [r.model_dump() for r in results]}


@router.post(
    "/generate-nav-files",
    dependencies=[Depends(_PERM_VIEW)],
)
async def generate_nav_files_endpoint(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    ids: Optional[str] = None,  # lista UUID separati da virgola
):
    if current_user.department not in ("IT", "ADMIN", "SUPERUSER"):
        raise HTTPException(403, "Accesso riservato alla gestione IT")

    nav_filters = [
        OperatorCodeRequest.is_evaded == True,
        OperatorCodeRequest.assigned_code.isnot(None),
    ]
    if ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        nav_filters.append(cast(OperatorCodeRequest.id, String).in_(id_list))

    result = await db.execute(
        select(OperatorCodeRequest).where(*nav_filters)
        .order_by(OperatorCodeRequest.store_number, OperatorCodeRequest.last_name)
    )
    to_export = result.scalars().all()

    if not to_export:
        return {"files": [], "count": 0}

    # Raggruppa per entity
    by_entity: dict = {}
    for req in to_export:
        entity = _entity_from_store(req.store_number)
        if not entity:
            continue
        by_entity.setdefault(entity, []).append({
            "assigned_code": req.assigned_code,
            "last_name": req.last_name,
            "first_name": req.first_name,
            "assigned_password": req.assigned_password,
        })

    if not by_entity:
        raise HTTPException(400, "Impossibile determinare l'entity per le richieste selezionate")

    # Genera file in memoria per ogni entity
    generated: list[dict] = []
    try:
        for entity, rows in by_entity.items():
            files = generate_nav_files(entity, rows)
            generated.extend(files)
    except Exception as e:
        raise HTTPException(500, f"Errore generazione file: {str(e)}")

    # Marca come esportate
    now = datetime.now(timezone.utc)
    for req in to_export:
        req.exported_at = now
    await db.commit()

    return {"files": generated, "count": len(to_export)}


@router.delete(
    "/clear",
    dependencies=[Depends(_PERM_VIEW)],
)
async def clear_operator_codes(db: AsyncSession = Depends(get_db)):
    await db.execute(delete(OperatorCode))
    await db.commit()
    return {"message": "Tabella operator_codes svuotata"}


@router.delete(
    "/pool/clear",
    dependencies=[Depends(_PERM_VIEW)],
)
async def clear_pool(db: AsyncSession = Depends(get_db)):
    await db.execute(delete(OperatorCodePool))
    await db.commit()
    return {"message": "Tabella operator_code_pool svuotata"}


@router.get(
    "/pool",
    dependencies=[Depends(_PERM_VIEW)],
)
async def list_pool(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(OperatorCodePool).order_by(OperatorCodePool.entity, OperatorCodePool.code)
    )
    rows = result.scalars().all()
    return [
        {
            "id": str(r.id),
            "entity": r.entity,
            "code": r.code,
            "full_name": r.full_name,
            "imported_at": r.imported_at.isoformat() if r.imported_at else None,
        }
        for r in rows
    ]


@router.post(
    "/pool/preview",
    response_model=PoolPreviewResponse,
    dependencies=[Depends(_PERM_VIEW)],
)
async def pool_preview(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File non valido: carica un file .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Impossibile leggere il file Excel")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    entity_col = next((col[k] for k in ["entity", "entità", "entita"] if k in col), None)
    code_col = next((col[k] for k in ["code", "codice", "cod"] if k in col), None)
    name_col = next((col[k] for k in ["name", "nome", "nominativo", "full_name", "nominative"] if k in col), None)

    if entity_col is None or code_col is None or name_col is None:
        raise HTTPException(
            400,
            f"Colonne non trovate. Attese: entity, code, name. Trovate: {list(col.keys())}",
        )

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_val = row[entity_col]
        code_val = row[code_col]
        name_val = row[name_col]

        if entity_val is None and code_val is None and name_val is None:
            continue

        entity_str = str(entity_val).strip().upper() if entity_val else None
        full_name = str(name_val).strip() if name_val else None

        if not entity_str or not code_val or not full_name:
            continue

        try:
            code_int = int(code_val)
        except (ValueError, TypeError):
            continue

        rows.append(PoolPreviewRow(entity=entity_str, code=code_int, full_name=full_name))

    return PoolPreviewResponse(rows=rows)


@router.post(
    "/pool/overwrite",
    dependencies=[Depends(_PERM_VIEW)],
)
async def pool_overwrite(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Sovrascrive ho.operator_code_pool con il contenuto di un file Excel."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File non valido: carica un file .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Impossibile leggere il file Excel")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    entity_col = next((col[k] for k in ["entity", "entità", "entita"] if k in col), None)
    code_col = next((col[k] for k in ["code", "codice", "cod"] if k in col), None)
    name_col = next((col[k] for k in ["name", "nome", "nominativo", "full_name", "nominative"] if k in col), None)

    if entity_col is None or code_col is None or name_col is None:
        raise HTTPException(
            400,
            f"Colonne non trovate. Attese: entity, code, name. Trovate: {list(col.keys())}",
        )

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_val = row[entity_col]
        code_val = row[code_col]
        name_val = row[name_col]

        if entity_val is None and code_val is None and name_val is None:
            continue

        entity_str = str(entity_val).strip().upper() if entity_val else None
        full_name = str(name_val).strip() if name_val else None

        if not entity_str or not code_val or not full_name:
            continue

        try:
            code_int = int(code_val)
        except (ValueError, TypeError):
            continue

        rows.append(OperatorCodePool(entity=entity_str, code=code_int, full_name=full_name))

    # Deduplica per (entity, code) — in caso di righe doppie nel file tiene l'ultima
    seen: dict = {}
    for r in rows:
        seen[(r.entity, r.code)] = r
    rows = list(seen.values())

    await db.execute(delete(OperatorCodePool))
    await db.commit()
    db.add_all(rows)
    await db.commit()
    return {"inserted": len(rows)}


@router.post(
    "/pool/import",
    dependencies=[Depends(_PERM_VIEW)],
)
async def pool_import(body: PoolImportPayload, db: AsyncSession = Depends(get_db)):
    existing_result = await db.execute(
        select(OperatorCodePool.entity, OperatorCodePool.code)
    )
    existing_keys = {(r.entity, r.code) for r in existing_result.all()}

    inserted = 0
    skipped = 0

    for row in body.rows:
        if (row.entity, row.code) in existing_keys:
            skipped += 1
            continue

        db.add(OperatorCodePool(entity=row.entity, code=row.code, full_name=row.full_name))
        existing_keys.add((row.entity, row.code))
        inserted += 1

    await db.commit()
    return {"inserted": inserted, "skipped": skipped}


@router.post(
    "/request",
    dependencies=[Depends(_PERM_VIEW)],
)
async def request_operator_code(
    data: OperatorCodeRequestPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Valida store_number
    store_result = await db.execute(
        select(Store).where(
            func.lower(Store.store_number) == data.store_number.lower(),
            Store.is_active == True,
        )
    )
    if not store_result.scalar_one_or_none():
        raise HTTPException(404, f"Negozio '{data.store_number}' non trovato")

    # Cerca match esatto in operator_codes
    exact_result = await db.execute(
        select(OperatorCode).where(
            func.lower(OperatorCode.first_name) == data.first_name.strip().lower(),
            func.lower(OperatorCode.last_name) == data.last_name.strip().lower(),
            OperatorCode.is_active == True,
        )
    )
    exact = exact_result.scalar_one_or_none()

    if exact:
        requester_name = creator_name = None
        if exact.requested_by:
            r = await db.execute(select(User).where(User.id == exact.requested_by))
            u = r.scalar_one_or_none()
            if u:
                requester_name = u.full_name
        if exact.created_by:
            r = await db.execute(select(User).where(User.id == exact.created_by))
            u = r.scalar_one_or_none()
            if u:
                creator_name = u.full_name
        return {"found": True, "pending": False, "code": _serialize_code(exact, requester_name, creator_name), "similar": [], "ticket_number": None}

    # Cerca match esatto in operator_code_requests (già in attesa)
    pending_result = await db.execute(
        select(OperatorCodeRequest).where(
            func.lower(OperatorCodeRequest.first_name) == data.first_name.strip().lower(),
            func.lower(OperatorCodeRequest.last_name) == data.last_name.strip().lower(),
        )
    )
    pending = pending_result.scalar_one_or_none()

    if pending:
        requester_name = None
        if pending.requested_by:
            r = await db.execute(select(User).where(User.id == pending.requested_by))
            u = r.scalar_one_or_none()
            if u:
                requester_name = u.full_name
        return {
            "found": False,
            "pending": True,
            "pending_request": {
                "first_name": pending.first_name,
                "last_name": pending.last_name,
                "store_number": pending.store_number,
                "start_date": pending.start_date.isoformat(),
                "created_at": pending.created_at.isoformat() if pending.created_at else None,
                "requester_name": requester_name,
            },
            "code": None,
            "similar": [],
            "ticket_number": None,
        }

    # Cerca candidati con stesso cognome (nomi simili / omonimia)
    similar_candidates = []
    if not data.force:
        sim_result = await db.execute(
            select(OperatorCode).where(
                func.lower(OperatorCode.last_name) == data.last_name.strip().lower(),
                OperatorCode.is_active == True,
            )
        )
        similar_candidates = sim_result.scalars().all()

    if similar_candidates:
        return {
            "found": False,
            "pending": False,
            "code": None,
            "similar": [_serialize_code(c) for c in similar_candidates],
            "ticket_number": None,
        }

    # Nessun match → inserisci in operator_code_requests e commita subito
    new_request = OperatorCodeRequest(
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        store_number=data.store_number.strip().upper(),
        start_date=data.start_date,
        requested_by=current_user.id,
    )
    db.add(new_request)
    await db.commit()

    # Crea ticket cumulativo solo se non esiste uno aperto (operazione separata)
    ticket_number = None
    existing_ticket = await _get_cumulative_ticket(db)
    if not existing_ticket:
        requester_name = current_user.full_name or current_user.username
        days_to_start = (data.start_date - date.today()).days
        priority = TicketPriority.HIGH if days_to_start < 2 else TicketPriority.MEDIUM

        ticket_data = TicketCreate(
            title="Richiesta Codice Operatore",
            description=(
                f"Nuove richieste di codice operatore in attesa di gestione.\n"
                f"Inoltrata da: {requester_name}\n\n"
                f"Operatori:\n  - {data.last_name} {data.first_name} — {data.store_number}"
            ),
            category_id=7,
            subcategory_id=17,
            priority=priority,
            requester_name=requester_name,
            requester_email=current_user.email or "",
            requester_phone="",
            teamviewer_code="",
        )
        try:
            ticket = await ticket_service.create_ticket(db, ticket_data, current_user)
            ticket_number = ticket.ticket_number
        except Exception:
            pass  # il new_request è già committato, il ticket è best-effort

    return {"found": False, "pending": False, "code": None, "similar": [], "ticket_number": ticket_number}
