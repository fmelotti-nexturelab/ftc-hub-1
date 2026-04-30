"""Gestione tabelle di riferimento del Converter + Assembly Engine.

Endpoint per tabella:
  GET  /api/items/converter/{table}/stats   → conteggio + ultima sincronizzazione
  GET  /api/items/converter/{table}         → lista paginata
  POST /api/items/converter/{table}/import  → carica Excel, sostituisce tutto

Assembly:
  GET  /api/items/converter/status          → stato tutte le tabelle
  POST /api/items/converter/assemble        → assembla item_master_it01
"""
import asyncio
import io
import json
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user, require_permission
from app.database import get_db
from app.models.auth import User
from app.services.items.assembly import run_assembly

router = APIRouter(prefix="/api/items/converter", tags=["Items - Converter Ref"])

_PERM        = require_permission("items_view")
_PERM_MANAGE = require_permission("items_view", need_manage=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _to_dt(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return None


async def _stats(db: AsyncSession, table: str) -> dict:
    row = (await db.execute(text(f"SELECT COUNT(*), MAX(synced_at) FROM ho.{table}"))).one()
    return {
        "count":    row[0],
        "last_sync": row[1].isoformat() if row[1] else None,
    }


async def _bulk_upsert(db: AsyncSession, sql: str, rows: list[dict], now, uid: str):
    for i in range(0, len(rows), 5000):
        await db.execute(text(sql), {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid})


# ── Config (path locali per script) ──────────────────────────────────────────

@router.get("/config", dependencies=[Depends(_PERM)])
async def get_converter_config(db: AsyncSession = Depends(get_db)):
    """Restituisce le impostazioni del Converter leggibili dagli script locali."""
    rows = (await db.execute(text(
        "SELECT setting_key, setting_value FROM ho.app_settings "
        "WHERE setting_key IN ('commercial_files_path')"
    ))).all()
    return {r[0]: r[1] for r in rows}


# ── IVA ───────────────────────────────────────────────────────────────────────

@router.get("/iva", dependencies=[Depends(_PERM)])
async def get_iva(db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text("SELECT vat_code, vat_pct FROM ho.item_iva ORDER BY vat_pct DESC"))).all()
    return [{"vat_code": r[0], "vat_pct": r[1]} for r in rows]


# ── KVI ───────────────────────────────────────────────────────────────────────

@router.get("/kvi/stats", dependencies=[Depends(_PERM)])
async def kvi_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_kvi")


@router.get("/kvi", dependencies=[Depends(_PERM)])
async def get_kvi(
    skip: int = 0, limit: int = 200, search: str = "",
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR item_name ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"SELECT item_no, item_name, type FROM ho.item_kvi {where} ORDER BY item_no LIMIT :lim OFFSET :skip"),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [{"item_no": r[0], "item_name": r[1], "type": r[2]} for r in rows]


@router.post("/kvi/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_kvi(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    if "KVI" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail='Foglio "KVI" non trovato nel file')
    ws = wb["KVI"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        item_no = _to_str(row[0])
        if not item_no:
            continue
        rows.append({"item_no": item_no, "item_name": _to_str(row[1]), "type": _to_str(row[2]) or "KVI"})
    wb.close()
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio KVI")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_kvi"))
    await db.execute(
        text("""
            INSERT INTO ho.item_kvi (item_no, item_name, type, synced_at, synced_by)
            SELECT r.item_no, r.item_name, r.type, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, item_name text, type text)
        """),
        {"rows": json.dumps(rows), "now": now, "uid": uid},
    )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── SB LIST ───────────────────────────────────────────────────────────────────

@router.get("/sb-list/stats", dependencies=[Depends(_PERM)])
async def sb_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_sb_list")


@router.get("/sb-list", dependencies=[Depends(_PERM)])
async def get_sb_list(
    skip: int = 0, limit: int = 200, search: str = "",
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR promo_name ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, promo_name, data_variazione, model_store_finale
            FROM ho.item_sb_list {where} ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [
        {"item_no": r[0], "promo_name": r[1],
         "data_variazione": r[2].isoformat() if r[2] else None,
         "model_store_finale": r[3]}
        for r in rows
    ]


@router.post("/sb-list/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_sb_list(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    if "SB LIST" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail='Foglio "SB LIST" non trovato nel file')
    ws = wb["SB LIST"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        item_no = _to_str(row[0])
        if not item_no:
            continue
        rows.append({
            "item_no": item_no, "promo_name": _to_str(row[1]),
            "data_variazione": _to_dt(row[2]), "model_store_finale": _to_str(row[4]),
        })
    wb.close()
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio SB LIST")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_sb_list"))
    await db.execute(
        text("""
            INSERT INTO ho.item_sb_list (item_no, promo_name, data_variazione, model_store_finale, synced_at, synced_by)
            SELECT r.item_no, r.promo_name,
                   CAST(NULLIF(r.data_variazione,'') AS timestamptz),
                   r.model_store_finale, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb))
              AS r(item_no text, promo_name text, data_variazione text, model_store_finale text)
        """),
        {"rows": json.dumps(rows), "now": now, "uid": uid},
    )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── CORE LIST ─────────────────────────────────────────────────────────────────

@router.get("/core-list/stats", dependencies=[Depends(_PERM)])
async def core_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_core_list")


@router.get("/core-list", dependencies=[Depends(_PERM)])
async def get_core_list(
    skip: int = 0, limit: int = 200, search: str = "",
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR ax_module_code ILIKE :q OR type ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, ax_module_code, type, type_original
            FROM ho.item_core_list {where} ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [{"item_no": r[0], "ax_module_code": r[1], "type": r[2], "type_original": r[3]} for r in rows]


@router.post("/core-list/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_core_list(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    if "CORE LIST" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail='Foglio "CORE LIST" non trovato nel file')
    ws = wb["CORE LIST"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        item_no = _to_str(row[0])
        if not item_no:
            continue
        rows.append({
            "item_no": item_no, "ax_module_code": _to_str(row[1]),
            "type": _to_str(row[4]), "type_original": _to_str(row[8]),
        })
    wb.close()
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio CORE LIST")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_core_list"))
    await db.execute(
        text("""
            INSERT INTO ho.item_core_list (item_no, ax_module_code, type, type_original, synced_at, synced_by)
            SELECT r.item_no, r.ax_module_code, r.type, r.type_original, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb))
              AS r(item_no text, ax_module_code text, type text, type_original text)
        """),
        {"rows": json.dumps(rows), "now": now, "uid": uid},
    )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── CAMPAGNE PROMO ────────────────────────────────────────────────────────────

@router.get("/campaigns/stats", dependencies=[Depends(_PERM)])
async def campaigns_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_campaigns_promo")


@router.get("/campaigns", dependencies=[Depends(_PERM)])
async def get_campaigns(
    skip: int = 0, limit: int = 200, search: str = "",
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR promo_name ILIKE :q OR status ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, promo_name, prezzo_attuale, prezzo_precedente,
                   data_variazione, fine_variazione, fine_promo,
                   type_item, type_after_promo, promo_in_cassa_pct, prezzo_netto, status
            FROM ho.item_campaigns_promo {where} ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [
        {
            "item_no": r[0], "promo_name": r[1],
            "prezzo_attuale": float(r[2]) if r[2] is not None else None,
            "prezzo_precedente": float(r[3]) if r[3] is not None else None,
            "data_variazione": r[4].isoformat() if r[4] else None,
            "fine_variazione": r[5].isoformat() if r[5] else None,
            "fine_promo": r[6].isoformat() if r[6] else None,
            "type_item": r[7], "type_after_promo": r[8],
            "promo_in_cassa_pct": float(r[9]) if r[9] is not None else None,
            "prezzo_netto": float(r[10]) if r[10] is not None else None,
            "status": r[11],
        }
        for r in rows
    ]


@router.post("/campaigns/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_campaigns(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if "CAMPAGNE" in s.upper() or "PROMO" in s.upper()), None
    )
    if not sheet_name:
        raise HTTPException(status_code=422, detail='Foglio campagne non trovato')
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        item_no = _to_str(row[0])
        if not item_no:
            continue
        rows.append({
            "item_no": item_no, "promo_name": _to_str(row[1]),
            "prezzo_attuale": _to_float(row[2]), "prezzo_precedente": _to_float(row[3]),
            "data_variazione": _to_dt(row[5]), "fine_variazione": _to_dt(row[6]),
            "fine_promo": _to_dt(row[7]), "type_item": _to_str(row[8]),
            "type_after_promo": _to_str(row[9]), "promo_in_cassa_pct": _to_float(row[10]),
            "prezzo_netto": _to_float(row[11]), "status": _to_str(row[12]),
        })
    wb.close()
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio campagne")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_campaigns_promo"))
    await db.execute(
        text("""
            INSERT INTO ho.item_campaigns_promo (
                item_no, promo_name, prezzo_attuale, prezzo_precedente,
                data_variazione, fine_variazione, fine_promo,
                type_item, type_after_promo, promo_in_cassa_pct, prezzo_netto, status,
                synced_at, synced_by
            )
            SELECT r.item_no, r.promo_name,
                   CAST(r.prezzo_attuale AS numeric), CAST(r.prezzo_precedente AS numeric),
                   CAST(NULLIF(r.data_variazione,'')  AS timestamptz),
                   CAST(NULLIF(r.fine_variazione,'')  AS timestamptz),
                   CAST(NULLIF(r.fine_promo,'')        AS timestamptz),
                   r.type_item, r.type_after_promo,
                   CAST(r.promo_in_cassa_pct AS numeric), CAST(r.prezzo_netto AS numeric),
                   r.status, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(
                item_no text, promo_name text,
                prezzo_attuale numeric, prezzo_precedente numeric,
                data_variazione text, fine_variazione text, fine_promo text,
                type_item text, type_after_promo text,
                promo_in_cassa_pct numeric, prezzo_netto numeric, status text
            )
        """),
        {"rows": json.dumps(rows), "now": now, "uid": uid},
    )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── RAW NAV (staging Appoggio) ────────────────────────────────────────────────

def _parse_raw_nav_ws(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        item_no = _to_str(row[0])
        if not item_no:
            continue
        barcode = None
        try:
            if row[8] is not None:
                barcode = int(float(str(row[8])))
        except (ValueError, TypeError):
            pass
        upc = None
        try:
            if row[10] is not None:
                upc = int(float(str(row[10])))
        except (ValueError, TypeError):
            pass
        rows.append({
            "item_no": item_no, "description": _to_str(row[1]) or "",
            "description_local": _to_str(row[2]), "warehouse": _to_str(row[3]),
            "last_cost": _to_float(row[4]), "unit_price": _to_float(row[5]),
            "item_cat": _to_str(row[6]), "net_weight": _to_float(row[7]),
            "barcode": barcode, "vat_code": _to_str(row[9]), "units_per_pack": upc,
        })
    return rows


@router.get("/raw-nav/stats", dependencies=[Depends(_PERM)])
async def raw_nav_stats(db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("SELECT COUNT(*), MAX(synced_at) FROM ho.item_raw_nav"))).one()
    return {"count": row[0], "last_sync": row[1].isoformat() if row[1] else None}


@router.get("/raw-nav", dependencies=[Depends(_PERM)])
async def get_raw_nav(
    skip: int = 0, limit: int = 200, search: str = "",
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR description ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, description, warehouse, unit_price, vat_code, units_per_pack
            FROM ho.item_raw_nav {where} ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [
        {"item_no": r[0], "description": r[1], "warehouse": r[2],
         "unit_price": float(r[3]) if r[3] else None,
         "vat_code": r[4], "units_per_pack": r[5]}
        for r in rows
    ]


@router.post("/raw-nav/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_raw_nav(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["Appoggio"] if "Appoggio" in wb.sheetnames else wb.active
        result = _parse_raw_nav_ws(ws)
        wb.close()
        return result

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato (attese 11 col: Nr., Descrizione, ...)")

    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_raw_nav"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_raw_nav (
                    item_no, description, description_local, warehouse,
                    last_cost, unit_price, item_cat, net_weight,
                    barcode, vat_code, units_per_pack, synced_at, synced_by
                )
                SELECT r.item_no, r.description, r.description_local, r.warehouse,
                       CAST(r.last_cost AS numeric), CAST(r.unit_price AS numeric),
                       r.item_cat, CAST(r.net_weight AS numeric),
                       CAST(r.barcode AS bigint), r.vat_code, CAST(r.units_per_pack AS integer),
                       :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(
                    item_no text, description text, description_local text, warehouse text,
                    last_cost numeric, unit_price numeric, item_cat text, net_weight numeric,
                    barcode bigint, vat_code text, units_per_pack integer
                )
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


class _RawNavTsvBody(BaseModel):
    tsv: str


def _parse_raw_nav_tsv(tsv_text: str) -> list[dict]:
    lines = tsv_text.strip().splitlines()
    if len(lines) < 2:
        return []
    rows = []
    for line in lines[1:]:
        cols = line.split("\t")
        while len(cols) < 11:
            cols.append(None)
        item_no = _to_str(cols[0])
        if not item_no:
            continue
        barcode = None
        try:
            if cols[8] is not None and str(cols[8]).strip():
                barcode = int(float(str(cols[8]).replace(",", ".")))
        except (ValueError, TypeError):
            pass
        upc = None
        try:
            if cols[10] is not None and str(cols[10]).strip():
                upc = int(float(str(cols[10]).replace(",", ".")))
        except (ValueError, TypeError):
            pass
        rows.append({
            "item_no": item_no, "description": _to_str(cols[1]) or "",
            "description_local": _to_str(cols[2]), "warehouse": _to_str(cols[3]),
            "last_cost": _to_float(cols[4]), "unit_price": _to_float(cols[5]),
            "item_cat": _to_str(cols[6]), "net_weight": _to_float(cols[7]),
            "barcode": barcode, "vat_code": _to_str(cols[9]), "units_per_pack": upc,
        })
    return rows


@router.post("/raw-nav/import-tsv", dependencies=[Depends(_PERM_MANAGE)])
async def import_raw_nav_tsv(
    body: _RawNavTsvBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = _parse_raw_nav_tsv(body.tsv)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato (attese 11 col: Nr., Descrizione, ...)")

    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_raw_nav"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_raw_nav (
                    item_no, description, description_local, warehouse,
                    last_cost, unit_price, item_cat, net_weight,
                    barcode, vat_code, units_per_pack, synced_at, synced_by
                )
                SELECT r.item_no, r.description, r.description_local, r.warehouse,
                       CAST(r.last_cost AS numeric), CAST(r.unit_price AS numeric),
                       r.item_cat, CAST(r.net_weight AS numeric),
                       CAST(r.barcode AS bigint), r.vat_code, CAST(r.units_per_pack AS integer),
                       :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(
                    item_no text, description text, description_local text, warehouse text,
                    last_cost numeric, unit_price numeric, item_cat text, net_weight numeric,
                    barcode bigint, vat_code text, units_per_pack integer
                )
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── PRICE (Country RP) ────────────────────────────────────────────────────────

@router.get("/price/stats", dependencies=[Depends(_PERM)])
async def price_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_price")


@router.get("/price", dependencies=[Depends(_PERM)])
async def list_price(
    search: str = "",
    limit: int = 300,
    skip: int = 0,
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"SELECT item_no, country_rp FROM ho.item_price {where} ORDER BY item_no LIMIT :lim OFFSET :skip"),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [{"item_no": r[0], "country_rp": float(r[1]) if r[1] is not None else None} for r in rows]


@router.post("/price/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_price(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["Price"] if "Price" in wb.sheetnames else wb.active
        result = []
        for row in ws.iter_rows(min_row=5, max_col=2, values_only=True):
            item_no = _to_str(row[0])
            crp = _to_float(row[1])
            if item_no and crp is not None:
                result.append({"item_no": item_no, "country_rp": crp})
        wb.close()
        return result

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio Price")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_price"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_price (item_no, country_rp, synced_at, synced_by)
                SELECT r.item_no, CAST(r.country_rp AS numeric), :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, country_rp numeric)
                ON CONFLICT (item_no) DO UPDATE
                    SET country_rp=EXCLUDED.country_rp, synced_at=EXCLUDED.synced_at
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


class _PriceJsonBody(BaseModel):
    items: list[dict]


@router.post("/price/import-json", dependencies=[Depends(_PERM_MANAGE)])
async def import_price_json(
    body: _PriceJsonBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = [
        {"item_no": str(r.get("item_no", "") or "").strip(), "country_rp": r.get("country_rp")}
        for r in body.items
        if r.get("item_no") and r.get("country_rp") is not None
    ]
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido ricevuto")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_price"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_price (item_no, country_rp, synced_at, synced_by)
                SELECT r.item_no, CAST(r.country_rp AS numeric), :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, country_rp numeric)
                ON CONFLICT (item_no) DO UPDATE
                    SET country_rp=EXCLUDED.country_rp, synced_at=EXCLUDED.synced_at
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── TRADUZIONI ────────────────────────────────────────────────────────────────

@router.get("/translations/stats", dependencies=[Depends(_PERM)])
async def translations_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_translations")


@router.post("/translations/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_translations(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["TRADUZIONI"] if "TRADUZIONI" in wb.sheetnames else wb.active
        result = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            item_no = _to_str(row[0])
            if item_no:
                result.append({"item_no": item_no, "d1": _to_str(row[1]), "d2": _to_str(row[2])})
        wb.close()
        return result

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato trovato nel foglio TRADUZIONI")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_translations"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_translations (item_no, descrizione1, descrizione2, synced_at, synced_by)
                SELECT r.item_no, r.d1, r.d2, :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, d1 text, d2 text)
                ON CONFLICT (item_no) DO UPDATE
                    SET descrizione1=EXCLUDED.descrizione1, descrizione2=EXCLUDED.descrizione2,
                        synced_at=EXCLUDED.synced_at
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── BOX SIZE ──────────────────────────────────────────────────────────────────

@router.get("/box-size/stats", dependencies=[Depends(_PERM)])
async def box_size_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_box_size")


@router.get("/box-size", dependencies=[Depends(_PERM)])
async def list_box_size(
    search: str = "",
    limit: int = 300,
    skip: int = 0,
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"SELECT item_no, box_size FROM ho.item_box_size {where} ORDER BY item_no LIMIT :lim OFFSET :skip"),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [{"item_no": r[0], "box_size": r[1]} for r in rows]


@router.post("/box-size/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_box_size(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["BOX SIZE"] if "BOX SIZE" in wb.sheetnames else wb.active
        result = {}
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            item_no = str(row[0]).strip() if row[0] is not None else ""
            if not item_no:
                continue
            try:
                box_val = int(float(str(row[2])))
                result[item_no] = max(box_val, 1)
            except (ValueError, TypeError):
                continue
        wb.close()
        return [{"item_no": k, "box_size": v} for k, v in result.items()]

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido nel foglio BOX SIZE")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_box_size"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_box_size (item_no, box_size, synced_at, synced_by)
                SELECT r.item_no, CAST(r.box_size AS integer), :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, box_size integer)
                ON CONFLICT (item_no) DO UPDATE SET box_size=EXCLUDED.box_size, synced_at=EXCLUDED.synced_at
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


class _BoxSizeJsonBody(BaseModel):
    items: list[dict]


@router.post("/box-size/import-json", dependencies=[Depends(_PERM_MANAGE)])
async def import_box_size_json(
    body: _BoxSizeJsonBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = []
    for r in body.items:
        item_no = str(r.get("item_no", "") or "").strip()
        box_size = r.get("box_size")
        if not item_no or box_size is None:
            continue
        try:
            rows.append({"item_no": item_no, "box_size": max(int(box_size), 1)})
        except (ValueError, TypeError):
            continue
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido ricevuto")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_box_size"))
    for i in range(0, len(rows), 5000):
        await db.execute(
            text("""
                INSERT INTO ho.item_box_size (item_no, box_size, synced_at, synced_by)
                SELECT r.item_no, CAST(r.box_size AS integer), :now, CAST(:uid AS uuid)
                FROM jsonb_to_recordset(CAST(:rows AS jsonb)) AS r(item_no text, box_size integer)
                ON CONFLICT (item_no) DO UPDATE SET box_size=EXCLUDED.box_size, synced_at=EXCLUDED.synced_at
            """),
            {"rows": json.dumps(rows[i:i+5000]), "now": now, "uid": uid},
        )
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── DISPLAY ───────────────────────────────────────────────────────────────────

_DISPLAY_KEYWORDS = [
    "Table", "Wall", "Behind the till", "Fridge", "Card wall",
    "Surprice bag area", "Bin", "Sales unit", "Candle wall",
]


def _extract_display_modulo(vm_module: str | None) -> str:
    if not vm_module or not vm_module.strip():
        return "ND"
    for kw in _DISPLAY_KEYWORDS:
        if kw in vm_module:
            return kw
    return "ND"


def _parse_display_rows(ws) -> list[dict]:
    result = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        vm_module = str(row[1]).strip() if row[1] is not None else ""
        try:
            flag = bool(int(float(str(row[2])))) if row[2] is not None else False
        except (ValueError, TypeError):
            flag = False
        result[item_no] = {
            "item_no":              item_no,
            "vm_module":            vm_module or None,
            "flag_hanging_display": flag,
            "modulo":               _extract_display_modulo(vm_module),
        }
    return list(result.values())


def _upsert_display(rows, now, uid):
    return (
        text("""
            INSERT INTO ho.item_display (item_no, vm_module, flag_hanging_display, modulo, synced_at, synced_by)
            SELECT r.item_no, r.vm_module, CAST(r.flag_hanging_display AS boolean),
                   r.modulo, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb))
              AS r(item_no text, vm_module text, flag_hanging_display boolean, modulo text)
            ON CONFLICT (item_no) DO UPDATE
                SET vm_module=EXCLUDED.vm_module, flag_hanging_display=EXCLUDED.flag_hanging_display,
                    modulo=EXCLUDED.modulo, synced_at=EXCLUDED.synced_at
        """),
        {"rows": None, "now": now, "uid": uid},
    )


@router.get("/display/stats", dependencies=[Depends(_PERM)])
async def display_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_display")


@router.get("/display", dependencies=[Depends(_PERM)])
async def list_display(
    search: str = "",
    limit: int = 300,
    skip: int = 0,
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR vm_module ILIKE :q OR modulo ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, vm_module, flag_hanging_display, modulo
            FROM ho.item_display {where}
            ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [
        {"item_no": r[0], "vm_module": r[1], "flag_hanging_display": r[2], "modulo": r[3]}
        for r in rows
    ]


@router.post("/display/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_display(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["Display"] if "Display" in wb.sheetnames else wb.active
        rows = _parse_display_rows(ws)
        wb.close()
        return rows

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido nel foglio Display")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_display"))
    for i in range(0, len(rows), 5000):
        stmt, params = _upsert_display(rows, now, uid)
        params["rows"] = json.dumps(rows[i:i+5000])
        await db.execute(stmt, params)
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


class _DisplayJsonBody(BaseModel):
    items: list[dict]


@router.post("/display/import-json", dependencies=[Depends(_PERM_MANAGE)])
async def import_display_json(
    body: _DisplayJsonBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = []
    for r in body.items:
        item_no = str(r.get("item_no", "") or "").strip()
        if not item_no:
            continue
        vm_module = str(r.get("vm_module") or "").strip()
        try:
            flag = bool(int(float(str(r.get("flag_hanging_display", 0) or 0))))
        except (ValueError, TypeError):
            flag = False
        rows.append({
            "item_no":              item_no,
            "vm_module":            vm_module or None,
            "flag_hanging_display": flag,
            "modulo":               _extract_display_modulo(vm_module),
        })
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido ricevuto")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_display"))
    for i in range(0, len(rows), 5000):
        stmt, params = _upsert_display(rows, now, uid)
        params["rows"] = json.dumps(rows[i:i+5000])
        await db.execute(stmt, params)
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── MASTER DATA BI ────────────────────────────────────────────────────────────

def _extract_item_type_bi(stat_name: str | None) -> str:
    if stat_name and str(stat_name).strip() == "Fixed":
        return "NA CORE"
    return "TAIL"


def _parse_master_bi_rows(ws) -> list[dict]:
    result = {}
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        barcode_ext = None
        try:
            if row[5] is not None:
                barcode_ext = int(float(str(row[5])))
        except (ValueError, TypeError):
            pass
        result[item_no] = {
            "item_no":      item_no,
            "category":     _to_str(row[3]),
            "subcategory":  _to_str(row[4]),
            "barcode_ext":  barcode_ext,
            "item_type_bi": _extract_item_type_bi(_to_str(row[2])),
        }
    return list(result.values())


def _upsert_master_bi(rows, now, uid):
    return (
        text("""
            INSERT INTO ho.item_master_bi (item_no, category, subcategory, barcode_ext, item_type_bi, synced_at, synced_by)
            SELECT r.item_no, r.category, r.subcategory, CAST(r.barcode_ext AS bigint),
                   r.item_type_bi, :now, CAST(:uid AS uuid)
            FROM jsonb_to_recordset(CAST(:rows AS jsonb))
              AS r(item_no text, category text, subcategory text, barcode_ext bigint, item_type_bi text)
            ON CONFLICT (item_no) DO UPDATE
                SET category=EXCLUDED.category, subcategory=EXCLUDED.subcategory,
                    barcode_ext=EXCLUDED.barcode_ext, item_type_bi=EXCLUDED.item_type_bi,
                    synced_at=EXCLUDED.synced_at
        """),
        {"rows": None, "now": now, "uid": uid},
    )


@router.get("/master-bi/stats", dependencies=[Depends(_PERM)])
async def master_bi_stats(db: AsyncSession = Depends(get_db)):
    return await _stats(db, "item_master_bi")


@router.get("/master-bi", dependencies=[Depends(_PERM)])
async def list_master_bi(
    search: str = "",
    limit: int = 300,
    skip: int = 0,
    db: AsyncSession = Depends(get_db),
):
    where = "WHERE item_no ILIKE :q OR category ILIKE :q OR subcategory ILIKE :q OR item_type_bi ILIKE :q" if search else ""
    q = f"%{search}%"
    rows = (await db.execute(
        text(f"""
            SELECT item_no, category, subcategory, barcode_ext, item_type_bi
            FROM ho.item_master_bi {where}
            ORDER BY item_no LIMIT :lim OFFSET :skip
        """),
        {"q": q, "lim": limit, "skip": skip},
    )).all()
    return [
        {"item_no": r[0], "category": r[1], "subcategory": r[2],
         "barcode_ext": r[3], "item_type_bi": r[4]}
        for r in rows
    ]


@router.post("/master-bi/import", dependencies=[Depends(_PERM_MANAGE)])
async def import_master_bi(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_bytes = await file.read()

    def _parse():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb["Master Data Bi"] if "Master Data Bi" in wb.sheetnames else wb.active
        rows = _parse_master_bi_rows(ws)
        wb.close()
        return rows

    rows = await asyncio.to_thread(_parse)
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido nel foglio Master Data Bi")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_master_bi"))
    for i in range(0, len(rows), 5000):
        stmt, params = _upsert_master_bi(rows, now, uid)
        params["rows"] = json.dumps(rows[i:i+5000])
        await db.execute(stmt, params)
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


class _MasterBiJsonBody(BaseModel):
    items: list[dict]


@router.post("/master-bi/import-json", dependencies=[Depends(_PERM_MANAGE)])
async def import_master_bi_json(
    body: _MasterBiJsonBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = []
    for r in body.items:
        item_no = str(r.get("item_no", "") or "").strip()
        if not item_no:
            continue
        barcode_ext = None
        try:
            if r.get("barcode_ext") is not None:
                barcode_ext = int(float(str(r["barcode_ext"])))
        except (ValueError, TypeError):
            pass
        rows.append({
            "item_no":      item_no,
            "category":     r.get("category"),
            "subcategory":  r.get("subcategory"),
            "barcode_ext":  barcode_ext,
            "item_type_bi": r.get("item_type_bi", "TAIL"),
        })
    if not rows:
        raise HTTPException(status_code=422, detail="Nessun dato valido ricevuto")
    now = datetime.now(timezone.utc)
    uid = str(current_user.id)
    await db.execute(text("TRUNCATE ho.item_master_bi"))
    for i in range(0, len(rows), 5000):
        stmt, params = _upsert_master_bi(rows, now, uid)
        params["rows"] = json.dumps(rows[i:i+5000])
        await db.execute(stmt, params)
    await db.commit()
    return {"synced": len(rows), "synced_at": now.isoformat()}


# ── STATUS GLOBALE + ASSEMBLY ─────────────────────────────────────────────────

_STATUS_TABLES = [
    ("raw-nav",      "item_raw_nav",         "Dati Raw NAV"),
    ("kvi",          "item_kvi",             "KVI"),
    ("campaigns",    "item_campaigns_promo",  "Campagne"),
    ("sb-list",      "item_sb_list",         "SB List"),
    ("core-list",    "item_core_list",        "Core List"),
    ("master-bi",    "item_master_bi",        "Master Data BI"),
    ("price",        "item_price",            "Price"),
    ("translations", "item_translations",     "Traduzioni"),
    ("box-size",     "item_box_size",         "Box Size"),
    ("display",      "item_display",          "Display"),
    ("iva",          "item_iva",              "IVA"),
]


@router.get("/status", dependencies=[Depends(_PERM)])
async def get_status(db: AsyncSession = Depends(get_db)):
    result = []
    for key, table, label in _STATUS_TABLES:
        row = (await db.execute(text(f"SELECT COUNT(*), MAX(synced_at) FROM ho.{table}"))).one()
        result.append({
            "key":      key,
            "label":    label,
            "count":    row[0],
            "last_sync": row[1].isoformat() if row[1] else None,
            "ok":       row[0] > 0,
        })
    return result


@router.post("/assemble", dependencies=[Depends(_PERM_MANAGE)])
async def assemble(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        return await run_assembly(db, str(current_user.id))
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
