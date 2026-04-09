import json
from datetime import date
from typing import Optional

import io
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user, require_permission
from app.database import get_db
from app.models.auth import User
from app.models.stock import StockItem, StockSession, StockStoreData
from app.services.stock_service import (
    cleanup_old_sessions,
    export_stock_xlsx,
    parse_stock_header,
    save_stock_data,
)

router = APIRouter(prefix="/api/stock", tags=["Stock"])

_PERM = require_permission("utilities_stock_nav")


# ---------------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------------

@router.post("/upload", dependencies=[Depends(_PERM)])
async def upload_stock(
    file: UploadFile = File(...),
    entity: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Il file deve essere un CSV")

    if entity not in ("IT01", "IT02", "IT03"):
        raise HTTPException(status_code=400, detail="Entity non valida — usare IT01, IT02 o IT03")

    content = await file.read()
    try:
        stock_date, store_codes = parse_stock_header(content, file.filename, entity)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    session = await save_stock_data(db, content, entity, file.filename, stock_date, store_codes, current_user.id)
    await cleanup_old_sessions(db, entity)

    return {
        "session_id": session.id,
        "entity": session.entity,
        "stock_date": str(session.stock_date),
        "total_items": session.total_items,
        "total_stores": session.total_stores,
    }


# ---------------------------------------------------------------------------
# Sessions
# ---------------------------------------------------------------------------

@router.get("/sessions", dependencies=[Depends(_PERM)])
async def get_sessions(
    entity: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    stmt = select(StockSession).order_by(StockSession.stock_date.desc())
    if entity:
        stmt = stmt.where(func.lower(StockSession.entity) == entity.lower())
    result = await db.execute(stmt)
    sessions = result.scalars().all()
    return [
        {
            "id": s.id,
            "entity": s.entity,
            "stock_date": str(s.stock_date),
            "filename": s.filename,
            "source": s.source,
            "total_items": s.total_items,
            "total_stores": s.total_stores,
            "store_codes": json.loads(s.store_codes_json) if s.store_codes_json else [],
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in sessions
    ]


@router.get("/sessions/{session_id}/stores", dependencies=[Depends(_PERM)])
async def get_session_stores(
    session_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Restituisce i codici negozio distinti per una sessione, letti direttamente dai dati."""
    # Verifica esistenza sessione
    sess_result = await db.execute(select(StockSession).where(StockSession.id == session_id))
    sess = sess_result.scalar_one_or_none()
    if not sess:
        raise HTTPException(status_code=404, detail="Sessione non trovata")

    # Legge sempre i negozi con dati reali (qty != 0) da StockStoreData
    result = await db.execute(
        select(StockStoreData.store_code)
        .join(StockItem, StockStoreData.item_id == StockItem.id)
        .where(StockItem.session_id == session_id)
        .distinct()
        .order_by(StockStoreData.store_code)
    )
    codes = [row[0] for row in result.all()]

    return {"session_id": session_id, "store_codes": codes}


# ---------------------------------------------------------------------------
# Items (article view)
# ---------------------------------------------------------------------------

@router.get("/sessions/{session_id}/items", dependencies=[Depends(_PERM)])
async def get_session_items(
    session_id: int,
    page: int = 1,
    page_size: int = 50,
    search: str = "",
    store_code: str = "",
    sort_by: str = "item_no",
    sort_dir: str = "asc",
    db: AsyncSession = Depends(get_db),
):
    # Verify session exists
    sess_result = await db.execute(select(StockSession).where(StockSession.id == session_id))
    sess = sess_result.scalar_one_or_none()
    if not sess:
        raise HTTPException(status_code=404, detail="Sessione non trovata")

    # Get store codes from session metadata (complete list, preserved from original CSV)
    store_codes = json.loads(sess.store_codes_json) if sess.store_codes_json else []

    # Base query for items
    stmt = select(StockItem).where(StockItem.session_id == session_id)

    if search:
        stmt = stmt.where(
            (StockItem.item_no.ilike(f"%{search}%")) |
            (StockItem.description.ilike(f"%{search}%"))
        )

    if store_code:
        # Only items with non-zero qty for that store
        stmt = stmt.where(
            StockItem.id.in_(
                select(StockStoreData.item_id).where(
                    StockStoreData.store_code == store_code,
                    StockStoreData.quantity != 0,
                )
            )
        )

    # Sort
    sort_col = StockItem.description if sort_by == "description" else StockItem.item_no
    stmt = stmt.order_by(sort_col.asc() if sort_dir == "asc" else sort_col.desc())

    # Count
    count_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = count_result.scalar()

    # Paginate
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    items_result = await db.execute(stmt)
    items = items_result.scalars().all()

    # Fetch store data for these items
    item_ids = [i.id for i in items]
    if item_ids:
        sd_result = await db.execute(
            select(StockStoreData).where(StockStoreData.item_id.in_(item_ids))
        )
        sd_rows = sd_result.scalars().all()
        # Group by item_id
        sd_map: dict[int, dict[str, int]] = {}
        for sd in sd_rows:
            sd_map.setdefault(sd.item_id, {})[sd.store_code] = sd.quantity
    else:
        sd_map = {}

    return {
        "items": [
            {
                "item_no": i.item_no,
                "description": i.description,
                "description_local": i.description_local,
                "adm_stock": i.adm_stock,
                "stores": sd_map.get(i.id, {}),
            }
            for i in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "stores": store_codes,
    }


# ---------------------------------------------------------------------------
# All items (no pagination — for Excel export)
# ---------------------------------------------------------------------------

@router.get("/sessions/{session_id}/all-items", dependencies=[Depends(_PERM)])
async def get_all_session_items(
    session_id: int,
    db: AsyncSession = Depends(get_db),
):
    sess_result = await db.execute(select(StockSession).where(StockSession.id == session_id))
    sess = sess_result.scalar_one_or_none()
    if not sess:
        raise HTTPException(status_code=404, detail="Sessione non trovata")

    store_codes = json.loads(sess.store_codes_json) if sess.store_codes_json else []

    # Fetch items + store data in one JOIN query — no IN clause, no param limit
    rows = await db.execute(
        text("""
            SELECT
                i.id, i.item_no, i.description, i.description_local, i.adm_stock,
                s.store_code, s.quantity
            FROM ho.stock_items i
            LEFT JOIN ho.stock_store_data s ON s.item_id = i.id
            WHERE i.session_id = :sid
            ORDER BY i.item_no, s.store_code
        """),
        {"sid": session_id},
    )

    items_map: dict[int, dict] = {}
    for row in rows.mappings():
        iid = row["id"]
        if iid not in items_map:
            items_map[iid] = {
                "item_no": row["item_no"],
                "description": row["description"],
                "description_local": row["description_local"],
                "adm_stock": row["adm_stock"],
                "stores": {},
            }
        if row["store_code"] is not None:
            items_map[iid]["stores"][row["store_code"]] = row["quantity"]

    return {
        "items": list(items_map.values()),
        "stores": store_codes,
        "total": len(items_map),
    }


# ---------------------------------------------------------------------------
# Store view
# ---------------------------------------------------------------------------

@router.get("/sessions/{session_id}/by-store/{store_code}", dependencies=[Depends(_PERM)])
async def get_session_by_store(
    session_id: int,
    store_code: str,
    page: int = 1,
    page_size: int = 50,
    search: str = "",
    hide_zero: bool = True,
    sort_by: str = "item_no",
    sort_dir: str = "asc",
    db: AsyncSession = Depends(get_db),
):
    sess_result = await db.execute(select(StockSession).where(StockSession.id == session_id))
    sess = sess_result.scalar_one_or_none()
    if not sess:
        raise HTTPException(status_code=404, detail="Sessione non trovata")

    stmt = (
        select(StockItem, StockStoreData.quantity)
        .join(StockStoreData, (StockStoreData.item_id == StockItem.id) & (StockStoreData.store_code == store_code))
        .where(StockItem.session_id == session_id)
    )

    if search:
        stmt = stmt.where(
            (StockItem.item_no.ilike(f"%{search}%")) |
            (StockItem.description.ilike(f"%{search}%"))
        )

    if hide_zero:
        stmt = stmt.where(StockStoreData.quantity != 0)

    sort_col = StockItem.description if sort_by == "description" else StockItem.item_no
    stmt = stmt.order_by(sort_col.asc() if sort_dir == "asc" else sort_col.desc())

    count_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = count_result.scalar()

    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    rows = await db.execute(stmt)

    return {
        "items": [
            {
                "item_no": item.item_no,
                "description": item.description,
                "description_local": item.description_local,
                "adm_stock": item.adm_stock,
                "quantity": qty,
            }
            for item, qty in rows.fetchall()
        ],
        "total": total,
        "store_code": store_code,
        "page": page,
        "page_size": page_size,
    }


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

@router.get("/sessions/{session_id}/export", dependencies=[Depends(_PERM)])
async def export_session(
    session_id: int,
    store_code: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        buf = await export_stock_xlsx(db, session_id, store_code)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    filename = f"stock_{session_id}"
    if store_code:
        filename += f"_{store_code}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# StockSplit CSV export
# ---------------------------------------------------------------------------

@router.get("/stocksplit", dependencies=[Depends(_PERM)])
async def stocksplit(
    stock_date: str,
    qty_filter: str = "positive",   # "all" | "positive" | "negative"
    include_adm: bool = False,
    db: AsyncSession = Depends(get_db),
):
    if qty_filter not in ("all", "positive", "negative"):
        raise HTTPException(status_code=400, detail="qty_filter deve essere 'all', 'positive' o 'negative'")

    try:
        parsed_date = date.fromisoformat(stock_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Data non valida. Formato atteso: YYYY-MM-DD")

    session_ids = []
    missing = []
    for entity in ("IT01", "IT02", "IT03"):
        result = await db.execute(
            select(StockSession).where(
                StockSession.entity == entity,
                StockSession.stock_date == parsed_date,
            )
        )
        sess = result.scalar_one_or_none()
        if sess:
            session_ids.append(sess.id)
        else:
            missing.append(entity)

    if missing:
        raise HTTPException(status_code=422, detail={"missing": missing})

    # ── Righe store ────────────────────────────────────────────────────────────
    store_stmt = (
        select(StockStoreData.store_code, StockItem.item_no, StockStoreData.quantity)
        .join(StockItem, StockItem.id == StockStoreData.item_id)
        .where(StockItem.session_id.in_(session_ids))
    )
    if qty_filter == "positive":
        store_stmt = store_stmt.where(StockStoreData.quantity > 0)
    elif qty_filter == "negative":
        store_stmt = store_stmt.where(StockStoreData.quantity < 0)
    store_stmt = store_stmt.order_by(StockStoreData.store_code.asc(), StockItem.item_no.asc())

    rows_result = await db.execute(store_stmt)
    rows = rows_result.fetchall()

    lines = ["STORE;Item No;QTY\n"]
    for row in rows:
        lines.append(f"{row.store_code};{row.item_no};{row.quantity}\n")

    # ── Righe ADM ─────────────────────────────────────────────────────────────
    if include_adm:
        adm_stmt = (
            select(StockItem.item_no, StockItem.adm_stock)
            .where(StockItem.session_id.in_(session_ids))
        )
        if qty_filter == "positive":
            adm_stmt = adm_stmt.where(StockItem.adm_stock > 0)
        elif qty_filter == "negative":
            adm_stmt = adm_stmt.where(StockItem.adm_stock < 0)
        adm_stmt = adm_stmt.order_by(StockItem.item_no.asc())

        adm_result = await db.execute(adm_stmt)
        for row in adm_result.fetchall():
            lines.append(f"ADM;{row.item_no};{row.adm_stock}\n")

    content = "".join(lines).encode("utf-8")
    date_str = stock_date.replace("-", "")
    filename = f"{date_str} STOCK SPLIT.csv"

    from fastapi import Response as FastAPIResponse
    return FastAPIResponse(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# ADM Extract — Excel con 3 fogli (ADM IT01 / ADM IT02 / ADM IT03)
# ---------------------------------------------------------------------------

@router.get("/adm-extract", dependencies=[Depends(_PERM)])
async def adm_extract(
    stock_date: str,
    db: AsyncSession = Depends(get_db),
):
    try:
        parsed_date = date.fromisoformat(stock_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Data non valida. Formato atteso: YYYY-MM-DD")

    session_ids = {}
    missing = []
    for entity in ("IT01", "IT02", "IT03"):
        result = await db.execute(
            select(StockSession).where(
                StockSession.entity == entity,
                StockSession.stock_date == parsed_date,
            )
        )
        sess = result.scalar_one_or_none()
        if sess:
            session_ids[entity] = sess.id
        else:
            missing.append(entity)

    if missing:
        raise HTTPException(status_code=422, detail={"missing": missing})

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuovi foglio vuoto default

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center")

    for entity in ("IT01", "IT02", "IT03"):
        ws = wb.create_sheet(title=f"ADM {entity}")

        # Intestazioni
        headers = ["No.", "Descrizione", "ADM"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Dati: tutti gli articoli della sessione, ordinati per item_no
        stmt = (
            select(StockItem.item_no, StockItem.description, StockItem.adm_stock)
            .where(StockItem.session_id == session_ids[entity])
            .order_by(StockItem.item_no.asc())
        )
        rows_result = await db.execute(stmt)
        rows = rows_result.fetchall()

        for r, row in enumerate(rows, 2):
            ws.cell(row=r, column=1, value=row.item_no)
            ws.cell(row=r, column=2, value=row.description)
            ws.cell(row=r, column=3, value=row.adm_stock)

        # Larghezze colonne
        ws.column_dimensions[get_column_letter(1)].width = 18
        ws.column_dimensions[get_column_letter(2)].width = 45
        ws.column_dimensions[get_column_letter(3)].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()

    date_str = stock_date.replace("-", "")
    filename = f"{date_str} ADM OneItaly.xlsx"

    from fastapi import Response as FastAPIResponse
    return FastAPIResponse(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Delete session
# ---------------------------------------------------------------------------

@router.delete("/sessions/{session_id}", dependencies=[Depends(_PERM)])
async def delete_session(
    session_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(StockSession).where(StockSession.id == session_id))
    sess = result.scalar_one_or_none()
    if not sess:
        raise HTTPException(status_code=404, detail="Sessione non trovata")
    await db.delete(sess)
    await db.commit()
    return {"message": "Sessione eliminata"}


# ---------------------------------------------------------------------------
# Stats (per dashboard card)
# ---------------------------------------------------------------------------

@router.get("/stats", dependencies=[Depends(_PERM)])
async def get_stats(db: AsyncSession = Depends(get_db)):
    stats = {}
    for entity in ("IT01", "IT02", "IT03"):
        result = await db.execute(
            select(StockSession)
            .where(StockSession.entity == entity)
            .order_by(StockSession.stock_date.desc())
            .limit(1)
        )
        sess = result.scalar_one_or_none()
        stats[entity] = {
            "stock_date": str(sess.stock_date) if sess else None,
            "total_items": sess.total_items if sess else 0,
            "total_stores": sess.total_stores if sess else 0,
        }
    return stats
