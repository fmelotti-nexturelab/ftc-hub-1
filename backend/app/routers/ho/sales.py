import logging
from datetime import date
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List
from openpyxl import load_workbook

from app.database import get_db
from app.models.auth import User
from app.models.ho import ExcludedStore
from app.schemas.ho import (
    ExcludedStoreCreate,
    ExcludedStoreResponse,
    SalesDataInput,
    SalesExportExcelRequest,
    SalesParseResponse,
)
from app.services.ho.sales import parse_tsv_nav
from app.services.app_settings_service import get_storage_path
from app.core.dependencies import require_ho, require_permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/ho/sales", tags=["HO - Sales"])

# ── Excluded Stores ────────────────────────────────────────────────────────────

@router.get(
    "/excluded-stores",
    response_model=List[ExcludedStoreResponse],
    dependencies=[Depends(require_permission("sales.view"))],
)
async def get_excluded_stores(
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExcludedStore).where(ExcludedStore.is_active == True)
    )
    return [ExcludedStoreResponse.model_validate(s) for s in result.scalars().all()]


@router.post(
    "/excluded-stores",
    response_model=ExcludedStoreResponse,
    dependencies=[Depends(require_permission("stores.exclude_manage"))],
)
async def add_excluded_store(
    data: ExcludedStoreCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_ho),
):
    store = ExcludedStore(**data.model_dump(), created_by=current_user.id)
    db.add(store)
    await db.commit()
    await db.refresh(store)
    return ExcludedStoreResponse.model_validate(store)


@router.delete(
    "/excluded-stores/{store_id}",
    dependencies=[Depends(require_permission("stores.exclude_manage"))],
)
async def remove_excluded_store(
    store_id: str,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExcludedStore).where(
            ExcludedStore.id == store_id,
            ExcludedStore.is_active == True,
        )
    )
    store = result.scalar_one_or_none()
    if not store:
        raise HTTPException(status_code=404, detail="Store non trovato")

    store.is_active = False
    await db.commit()
    return {"message": "Store rimosso dalla lista esclusi"}


@router.post(
    "/parse",
    response_model=SalesParseResponse,
    dependencies=[Depends(require_permission("sales.import"))],
)
async def parse_sales_data(
    data: SalesDataInput,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExcludedStore).where(ExcludedStore.is_active == True)
    )
    excluded_rows = result.scalars().all()
    excluded = {s.store_code.upper(): s.reason.value for s in excluded_rows}

    return SalesParseResponse(
        it01=parse_tsv_nav(data.raw_tsv_it01, "IT01", excluded) if data.raw_tsv_it01 else None,
        it02=parse_tsv_nav(data.raw_tsv_it02, "IT02", excluded) if data.raw_tsv_it02 else None,
        it03=parse_tsv_nav(data.raw_tsv_it03, "IT03", excluded) if data.raw_tsv_it03 else None,
    )


@router.post(
    "/export-excel",
    dependencies=[Depends(require_permission("sales.export"))],
)
async def export_sales_to_excel(
    data: SalesExportExcelRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Aggiorna il file RetailSalesAnalysis.xlsx con i dati di vendita.
    - Foglio 1: colonna A = codici negozio, riga 1 = date
    - Aggiunge la colonna con la data di analisi se non esiste
    - Scrive i valori per ogni negozio
    """
    try:
        storage_root = await get_storage_path(db)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    # Traduzione path Windows → Linux per container Docker
    # es. "F:\FTC_HUB_Archivio" → "/mnt/f/FTC_HUB_Archivio"
    import re
    drive_match = re.match(r'^([A-Za-z]):[/\\](.*)$', storage_root)
    if drive_match:
        storage_root = f"/mnt/{drive_match.group(1).lower()}/{drive_match.group(2)}"
    storage_root = storage_root.replace("\\", "/")

    year = f"20{data.analysis_date.split('.')[-1]}" if len(data.analysis_date.split('.')[-1]) == 2 else data.analysis_date.split('.')[-1]
    file_path = Path(storage_root) / "03_RetailSalesAnalysis" / year / "RetailSalesAnalysis.xlsx"

    if not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"File non trovato: {file_path}",
        )

    try:
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]

        # Trova o crea la colonna per la data di analisi
        date_col = None
        for col_idx in range(2, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col_idx).value
            if cell_val is not None and str(cell_val).strip() == data.analysis_date:
                date_col = col_idx
                break

        if date_col is None:
            date_col = ws.max_column + 1
            ws.cell(row=1, column=date_col, value=data.analysis_date)
            logger.info("Aggiunta colonna data %s alla posizione %d", data.analysis_date, date_col)

        # Mappa codice negozio → riga
        store_row_map = {}
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val:
                store_row_map[str(cell_val).strip().upper()] = row_idx

        # Scrivi i valori
        written = 0
        for item in data.stores:
            row_idx = store_row_map.get(item.store_code.upper())
            if row_idx:
                ws.cell(row=row_idx, column=date_col, value=item.value)
                written += 1

        wb.save(file_path)
        logger.info("Salvato %s: %d negozi aggiornati per data %s", file_path, written, data.analysis_date)

        return {"message": f"{written} negozi aggiornati", "file": str(file_path)}

    except PermissionError:
        raise HTTPException(status_code=500, detail="File Excel aperto da un altro programma. Chiudilo e riprova.")
    except Exception as e:
        logger.error("Errore aggiornamento Excel: %s", e)
        raise HTTPException(status_code=500, detail=f"Errore aggiornamento Excel: {str(e)}")

