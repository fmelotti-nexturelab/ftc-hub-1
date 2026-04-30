"""
sales_sync.py — Sincronizzazione vendite L2W verso FTC HUB

Uso:
    python sales_sync.py            # refresh + sync
    python sales_sync.py --no-excel # solo invio dati (file già aggiornato)

Configurazione in sales_sync.cfg (stesso folder):
    [ftchub]
    url  = http://localhost:8000
    user = admin@ftchub.it
    pass = password

    [sales]
    master_path = C:\\Users\\fmelo\\Zebra A S\\...\\SALES X WEEK MASTER.xlsx

Task Scheduler: punta a questo script, frequenza giornaliera mattina.
"""

import configparser
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import requests
import win32com.client


# ── Config ────────────────────────────────────────────────────────────────────

_SCRIPT_DIR = Path(__file__).parent
_CFG_PATH   = _SCRIPT_DIR / "sales_sync.cfg"

_cfg = configparser.ConfigParser()
_cfg.read(_CFG_PATH, encoding="utf-8")

_USERNAME = os.environ.get("USERNAME", "fmelo")
_DEFAULT_MASTER = (
    rf"C:\Users\{_USERNAME}\Zebra A S\One Italy Commercial - Files"
    r"\06 - MASTER data tools & services\01 - MasterTables\SALES X WEEK MASTER.xlsx"
)

BACKEND_URL   = _cfg.get("ftchub", "url",  fallback="http://localhost:8000").rstrip("/")
BACKEND_USER  = _cfg.get("ftchub", "user", fallback="")
BACKEND_PASS  = _cfg.get("ftchub", "pass", fallback="")
_master_cfg   = _cfg.get("sales", "master_path", fallback="").strip()
MASTER_PATH   = Path(_master_cfg if _master_cfg else _DEFAULT_MASTER)


# ── Logging ───────────────────────────────────────────────────────────────────

_LOG_PATH = _SCRIPT_DIR / "sales_sync.log"

def log(msg: str) -> None:
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line, flush=True)
    with open(_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ── Normalizzazione codice negozio ───────────────────────────────────────────

def _normalize_store_code(code: str) -> str:
    """Converte IT01001 (formato Excel) → IT101 (formato FTC HUB).
    Formula: entity * 100 + store  (es. entity=1, store=1 → IT101)
    """
    if len(code) == 7 and code.startswith("IT") and code[2:4].isdigit() and code[4:7].isdigit():
        entity = int(code[2:4])
        store  = int(code[4:7])
        return f"IT{entity * 100 + store}"
    return code


# ── Settimane L2W ─────────────────────────────────────────────────────────────

def get_l2w_weeks() -> tuple[list[str], str, str]:
    """Ritorna (lista, week_from, week_to) in formato YY.WW — settimana corrente + 2 precedenti."""
    weeks = []
    for i in range(3):
        d = date.today() - timedelta(weeks=i)
        y, w, _ = d.isocalendar()
        weeks.append(f"{str(y)[2:]}.{w:02d}")
    return weeks, weeks[-1], weeks[0]


# ── Excel: refresh pivot ──────────────────────────────────────────────────────

def prepare_excel(file_path: Path, update_weeks: bool, weeks_to_show: list[str]) -> None:
    """Apre Excel, aggiorna filtri (se lunedì), refresha, salva e chiude."""
    log("Apertura Excel...")
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(file_path))
        ws = wb.Sheets("Sellout")
        pt = ws.PivotTables(1)

        if update_weeks:
            log(f"Aggiornamento filtri Year Week → {weeks_to_show}")
            pf = pt.PivotFields("Year Week")
            weeks_set = set(weeks_to_show)
            # Prima abilita le settimane target (evita il caso "almeno 1 visibile")
            for item in pf.PivotItems():
                if str(item.Value) in weeks_set:
                    try:
                        item.Visible = True
                    except Exception:
                        pass
            # Poi disabilita tutte le altre
            for item in pf.PivotItems():
                if str(item.Value) not in weeks_set:
                    try:
                        item.Visible = False
                    except Exception:
                        pass

        log("Impostazione query sincrona...")
        for conn in wb.Connections:
            try:
                conn.OLEDBConnection.BackgroundQuery = False
            except Exception:
                try:
                    conn.ODBCConnection.BackgroundQuery = False
                except Exception:
                    pass

        log("Refresh in corso (potrebbe richiedere qualche minuto)...")
        wb.RefreshAll()
        xl.CalculateUntilAsyncQueriesDone()
        wb.Save()
        log("Refresh completato, file salvato.")

    except Exception as e:
        log(f"ERRORE in prepare_excel: {e}")
        raise
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        try:
            xl.Quit()
        except Exception:
            pass


# ── Excel: lettura dati Sellout ───────────────────────────────────────────────

def read_sellout(file_path: Path) -> list[dict]:
    """Legge il foglio Sellout con openpyxl e ritorna lista {item_no, store_code, qty_sold}."""
    log("Lettura dati Sellout (caricamento in memoria)...")
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    if "Sellout" not in wb.sheetnames:
        wb.close()
        raise RuntimeError('Foglio "Sellout" non trovato nel file')

    ws = wb["Sellout"]

    # Carica tutte le righe in memoria in una sola passata (molto più veloce di cell-by-cell)
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    log(f"Caricate {len(all_rows)} righe.")

    # Trova la riga header con "Etichette di riga"
    header_idx = None
    for i, row in enumerate(all_rows):
        for cell in row:
            if cell and "etichette di riga" in str(cell).lower():
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is None:
        raise RuntimeError("Intestazione pivot non trovata nel foglio Sellout")

    header = all_rows[header_idx]

    # Legge i codici negozio dalla riga header (da colonna 1 in poi)
    stores: list[tuple[int, str]] = []
    for col_idx in range(1, len(header)):
        val = header[col_idx]
        if val is None:
            break
        val_str = str(val).strip()
        if not val_str:
            continue
        low = val_str.lower()
        if low.startswith("totale") or low.startswith("grand"):
            break
        stores.append((col_idx, val_str))

    log(f"Trovati {len(stores)} negozi nell'intestazione.")

    # Legge i dati
    records: list[dict] = []
    for row in all_rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        item_no_str = str(row[0]).strip()
        if not item_no_str:
            continue
        low = item_no_str.lower()
        if low.startswith("totale") or low.startswith("grand"):
            break
        for col_idx, store_code in stores:
            if col_idx >= len(row):
                continue
            qty = row[col_idx]
            if qty is not None and isinstance(qty, (int, float)) and qty > 0:
                records.append({
                    "item_no":    item_no_str,
                    "store_code": _normalize_store_code(store_code),
                    "qty_sold":   int(qty),
                })

    log(f"Letti {len(records)} record.")
    return records


# ── Backend: login + sync ─────────────────────────────────────────────────────

def backend_login(session: requests.Session) -> None:
    if not BACKEND_USER or not BACKEND_PASS:
        raise RuntimeError(
            "Credenziali backend mancanti — configura [ftchub] user/pass in sales_sync.cfg"
        )
    log(f"Login backend ({BACKEND_USER})...")
    r = session.post(
        f"{BACKEND_URL}/api/auth/login",
        json={"username": BACKEND_USER, "password": BACKEND_PASS},
        timeout=30,
    )
    r.raise_for_status()
    token = r.json().get("access_token")
    if not token:
        raise RuntimeError(f"Login fallito: {r.text}")
    session.headers["Authorization"] = f"Bearer {token}"
    log("Login OK.")


def backend_sync(session: requests.Session, records: list[dict], week_from: str, week_to: str) -> None:
    log(f"Invio {len(records)} record al backend (settimane {week_from} → {week_to})...")
    r = session.post(
        f"{BACKEND_URL}/api/items/sales-l2w/sync",
        json={"items": records, "week_from": week_from, "week_to": week_to},
        timeout=120,
    )
    r.raise_for_status()
    log(f"Sync completato: {r.json()}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    no_excel = "--no-excel" in sys.argv

    log(f"Percorso file: {MASTER_PATH}")
    if not MASTER_PATH.exists():
        log(f"ERRORE: file non trovato: {MASTER_PATH}")
        log("Verifica il percorso in sales_sync.cfg -> [sales] master_path")
        sys.exit(1)

    is_monday = date.today().weekday() == 0
    weeks, week_from, week_to = get_l2w_weeks()
    log(f"Settimane L2W: {week_from} -> {week_to}  (lista: {weeks})")

    if not no_excel:
        log(f"E' lunedi: {is_monday} - filtri {'verranno aggiornati' if is_monday else 'invariati'}")
        prepare_excel(MASTER_PATH, is_monday, weeks)
    else:
        log("--no-excel: salto refresh Excel.")

    records = read_sellout(MASTER_PATH)
    if not records:
        log("ATTENZIONE: nessun record trovato - controlla che il pivot abbia dati.")
        sys.exit(1)

    session = requests.Session()
    backend_login(session)
    backend_sync(session, records, week_from, week_to)
    log("Completato.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        log(f"ERRORE FATALE: {e}")
        log(traceback.format_exc())
        sys.exit(1)
