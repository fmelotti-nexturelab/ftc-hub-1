"""ftchub-nav-agent — agent locale per operazioni desktop (Excel COM, RDP, ExpoList)
Porta: 19999
"""

import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
import urllib3
import uvicorn
import win32com.client
import win32con
import win32gui
import win32process
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Percorsi file ─────────────────────────────────────────────────────────────
# Le directory base possono essere sovrascritte tramite variabili d'ambiente.
_EXPO_BASE = Path(
    os.environ.get(
        "EXPO_LIST_BASE_DIR",
        r"C:\Users\fmelo\Zebra A S\One Italy Stores - Files",
    )
)
EXPO_LIST_PATH      = _EXPO_BASE / r"00 - Estrazioni\97 - Service\01 - Tables\tbl_ExpoList.xlsm"
ECO_LIST_PATH       = _EXPO_BASE / r"00 - Estrazioni\97 - Service\01 - Tables\tbl_ECO.xlsx"
ECCEZIONI_LIST_PATH = _EXPO_BASE / r"00 - Estrazioni\97 - Service\01 - Tables\tbl_Eccezioni.xlsm"
KGL_LIST_PATH       = _EXPO_BASE / r"00 - Estrazioni\97 - Service\01 - Tables\tbl_KGL.xlsm"


# ── Modelli request ──────────────────────────────────────────────────────────

class OpenRequest(BaseModel):
    path: str

class OpenConverterRequest(BaseModel):
    url: str
    filenameMatch: str

class PasteRequest(BaseModel):
    filenameMatch: str
    sheetName: str
    rows: list[list]


# ── Helper: trova workbook Excel aperto per nome ─────────────────────────────

def _find_workbook(xl, filename_match: str):
    for wb in xl.Workbooks:
        if filename_match.lower() in wb.Name.lower():
            return wb
    return None


# ── Endpoint esistenti ───────────────────────────────────────────────────────

@app.get("/ping")
def ping():
    return {"status": "ok"}


@app.get("/version")
def version():
    return {"version": "2", "type": "ftchub-nav-agent"}


@app.post("/restart")
def restart():
    """Riavvia il processo agente (usato dal frontend per applicare aggiornamenti)."""
    def _do():
        time.sleep(0.4)
        os.execv(sys.executable, [sys.executable] + sys.argv)
    threading.Thread(target=_do, daemon=True).start()
    return {"ok": True}


@app.post("/open")
def open_rdp(req: OpenRequest):
    try:
        subprocess.Popen(["mstsc.exe", req.path])
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/focus")
def focus_rdp():
    def _enum(hwnd, _):
        if win32gui.IsWindowVisible(hwnd):
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
            try:
                import psutil
                proc = psutil.Process(pid)
                if "mstsc" in proc.name().lower():
                    win32gui.SetForegroundWindow(hwnd)
            except Exception:
                pass
    win32gui.EnumWindows(_enum, None)
    return {"ok": True}


@app.post("/kill")
def kill_sessions():
    killed = 0
    def _enum(hwnd, _):
        nonlocal killed
        if win32gui.IsWindowVisible(hwnd):
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
            try:
                import psutil
                proc = psutil.Process(pid)
                if "mstsc" in proc.name().lower():
                    proc.kill()
                    killed += 1
            except Exception:
                pass
    win32gui.EnumWindows(_enum, None)
    return {"killed": killed}


@app.post("/browse-folder")
def browse_folder():
    """Apre un FolderBrowserDialog Windows nativo tramite PowerShell."""
    ps_script = (
        "Add-Type -AssemblyName System.Windows.Forms;"
        "$dlg = New-Object System.Windows.Forms.FolderBrowserDialog;"
        "$dlg.Description = 'Seleziona la cartella';"
        "$dlg.ShowNewFolderButton = $false;"
        "$r = $dlg.ShowDialog();"
        "if ($r -eq 'OK') { Write-Output $dlg.SelectedPath }"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NonInteractive", "-Command", ps_script],
            capture_output=True, text=True, timeout=60,
        )
        path = result.stdout.strip()
        if path:
            return {"ok": True, "path": path}
        return {"ok": False, "cancelled": True}
    except subprocess.TimeoutExpired:
        return {"ok": False, "cancelled": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/open-converter")
def open_converter(req: OpenConverterRequest):
    """Apre il file Convertitore in Excel desktop via protocollo ms-excel.
    Se già aperto, porta la finestra in primo piano.
    Fa polling finché Excel non ha il file attivo (timeout 60s).
    """
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True

    wb = _find_workbook(xl, req.filenameMatch)
    if not wb:
        os.startfile(f"ms-excel:ofe|u|{req.url}")
        deadline = time.time() + 60
        while time.time() < deadline:
            time.sleep(1)
            wb = _find_workbook(xl, req.filenameMatch)
            if wb:
                break
        if not wb:
            raise HTTPException(status_code=504, detail="Timeout: Excel non ha aperto il file entro 60s")

    # Porta in primo piano
    try:
        hwnd = xl.Hwnd
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.SetForegroundWindow(hwnd)
    except Exception:
        pass

    return {"ok": True, "workbook": wb.Name}


@app.post("/paste-to-converter")
def paste_to_converter(req: PasteRequest):
    """Scrive rows nel foglio sheetName del workbook che contiene filenameMatch."""
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        raise HTTPException(status_code=404, detail="Excel non è aperto")

    wb = _find_workbook(xl, req.filenameMatch)
    if not wb:
        raise HTTPException(status_code=404, detail=f'Workbook "{req.filenameMatch}" non trovato in Excel')

    try:
        ws = wb.Sheets(req.sheetName)
    except Exception:
        raise HTTPException(status_code=422, detail=f'Foglio "{req.sheetName}" non trovato')

    ws.Cells.ClearContents()
    for r_idx, row in enumerate(req.rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.Cells(r_idx, c_idx).Value = val

    return {"ok": True, "rows_written": len(req.rows)}


# ── Nuovi endpoint ExpoList ──────────────────────────────────────────────────

@app.get("/expo-list-mtime")
def expo_list_mtime():
    """Ritorna la data di ultima modifica di tbl_ExpoList.xlsm."""
    if not EXPO_LIST_PATH.exists():
        return {"exists": False, "mtime": None, "path": str(EXPO_LIST_PATH)}
    mtime = datetime.fromtimestamp(EXPO_LIST_PATH.stat().st_mtime).isoformat()
    return {"exists": True, "mtime": mtime, "path": str(EXPO_LIST_PATH)}


@app.get("/expo-list-data")
def expo_list_data():
    """Legge il foglio EXPO da tbl_ExpoList.xlsm e ritorna lista {item_no, expo_type}."""
    if not EXPO_LIST_PATH.exists():
        raise HTTPException(status_code=404, detail=f"File non trovato: {EXPO_LIST_PATH}")

    try:
        wb = openpyxl.load_workbook(str(EXPO_LIST_PATH), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore apertura file: {e}")

    if "EXPO" not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=422, detail='Foglio "EXPO" non trovato nel file')

    ws = wb["EXPO"]
    valid = {"TABLE", "WALL", "BUCKET"}
    items = []

    # Riga 1: metadato Power Query → skip
    # Riga 2: header (Item Number, EXPO) → skip
    # Dati da riga 3
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        item_no  = row[0]
        expo_raw = row[1]
        if not item_no or not expo_raw:
            continue
        item_no_str = str(item_no).strip()
        expo_str    = str(expo_raw).strip().upper()
        if item_no_str and expo_str in valid:
            items.append({"item_no": item_no_str, "expo_type": expo_str})

    wb.close()
    return {"items": items, "count": len(items)}


# ── Endpoint ECO List ────────────────────────────────────────────────────────

@app.get("/eco-list-mtime")
def eco_list_mtime():
    """Ritorna la data di ultima modifica di tbl_ECO.xlsx."""
    if not ECO_LIST_PATH.exists():
        return {"exists": False, "mtime": None, "path": str(ECO_LIST_PATH)}
    mtime = datetime.fromtimestamp(ECO_LIST_PATH.stat().st_mtime).isoformat()
    return {"exists": True, "mtime": mtime, "path": str(ECO_LIST_PATH)}


@app.get("/eco-list-data")
def eco_list_data():
    """Legge il foglio Foglio1 da tbl_ECO.xlsx e ritorna lista di item_no ECO."""
    if not ECO_LIST_PATH.exists():
        raise HTTPException(status_code=404, detail=f"File non trovato: {ECO_LIST_PATH}")

    try:
        wb = openpyxl.load_workbook(str(ECO_LIST_PATH), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore apertura file: {e}")

    # Il file ha un solo foglio (Foglio1); usa il primo disponibile
    ws = wb.active
    items = []

    # Riga 1: header (Zebra, Eco) → skip; dati da riga 2
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        item_no = row[0]
        if not item_no:
            continue
        item_no_str = str(item_no).strip()
        if item_no_str:
            items.append(item_no_str)

    wb.close()
    return {"items": items, "count": len(items)}


# ── Endpoint Eccezioni / BestSeller ─────────────────────────────────────────

@app.get("/eccezioni-mtime")
def eccezioni_mtime():
    """Ritorna la data di ultima modifica di tbl_Eccezioni.xlsm."""
    if not ECCEZIONI_LIST_PATH.exists():
        return {"exists": False, "mtime": None, "path": str(ECCEZIONI_LIST_PATH)}
    mtime = datetime.fromtimestamp(ECCEZIONI_LIST_PATH.stat().st_mtime).isoformat()
    return {"exists": True, "mtime": mtime, "path": str(ECCEZIONI_LIST_PATH)}


@app.get("/eccezioni-data")
def eccezioni_data():
    """Legge i fogli ECCEZIONI e BEST SELLER da tbl_Eccezioni.xlsm."""
    if not ECCEZIONI_LIST_PATH.exists():
        raise HTTPException(status_code=404, detail=f"File non trovato: {ECCEZIONI_LIST_PATH}")

    try:
        wb = openpyxl.load_workbook(str(ECCEZIONI_LIST_PATH), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore apertura file: {e}")

    # ── Foglio ECCEZIONI ──────────────────────────────────────────────────────
    # Riga 1: metadato Power Query → skip
    # Riga 2: header (ZEBRA, DESCRIZIONE, 1 PREZZO, 2 PREZZO, SCONTO, TESTO...) → skip
    # Dati da riga 3; colonne A-F (indici 0-5)
    eccezioni = []
    if "ECCEZIONI" in wb.sheetnames:
        ws_ecc = wb["ECCEZIONI"]
        for row in ws_ecc.iter_rows(min_row=3, max_col=12, values_only=True):
            zebra = row[0]
            if not zebra:
                continue
            zebra_str = str(zebra).strip()
            if not zebra_str:
                continue

            def _sv(v):
                return str(v).strip() if v is not None and str(v).strip() not in ("", "0.0", "None") else None

            def _fv(v):
                if v is None:
                    return None
                try:
                    return float(str(v).replace(",", "."))
                except (ValueError, TypeError):
                    return None

            eccezioni.append({
                "zebra":        zebra_str,
                "descrizione":  _sv(row[1]),
                "prezzo_1":     _fv(row[2]),
                "prezzo_2":     _fv(row[3]),
                "sconto":       _sv(row[4]),
                "testo_prezzo": _sv(row[5]),
                "categoria":    _sv(row[6]),
                "eccezione":    _sv(row[7]),
                "testo_prezzo2":_sv(row[8]),
                "col11":        _sv(row[9]),
                "col12":        _sv(row[10]),
            })

    # ── Foglio BEST SELLER ────────────────────────────────────────────────────
    # Riga 1: header (ITEM) → skip; dati da riga 2
    bestseller = []
    if "BEST SELLER" in wb.sheetnames:
        ws_bs = wb["BEST SELLER"]
        for row in ws_bs.iter_rows(min_row=2, max_col=1, values_only=True):
            item_no = row[0]
            if not item_no:
                continue
            item_no_str = str(item_no).strip()
            if item_no_str:
                bestseller.append({"item_no": item_no_str})

    wb.close()
    return {
        "eccezioni":  eccezioni,
        "bestseller": bestseller,
        "eccezioni_count":  len(eccezioni),
        "bestseller_count": len(bestseller),
    }


# ── Endpoint KGL (pesi corretti) ─────────────────────────────────────────────

@app.get("/kgl-mtime")
def kgl_mtime():
    """Ritorna la data di ultima modifica di tbl_KGL.xlsm."""
    if not KGL_LIST_PATH.exists():
        return {"exists": False, "mtime": None, "path": str(KGL_LIST_PATH)}
    mtime = datetime.fromtimestamp(KGL_LIST_PATH.stat().st_mtime).isoformat()
    return {"exists": True, "mtime": mtime, "path": str(KGL_LIST_PATH)}


@app.get("/kgl-data")
def kgl_data():
    """Legge il foglio KG-L da tbl_KGL.xlsm (col A = Nr., col D = PESO CORRETTO).

    Struttura: riga 1 = metadato Power Query, riga 2 = header, dati da riga 3.
    Solo righe con PESO CORRETTO numerico valido (> 0) vengono incluse.
    """
    if not KGL_LIST_PATH.exists():
        raise HTTPException(status_code=404, detail=f"File non trovato: {KGL_LIST_PATH}")

    try:
        wb = openpyxl.load_workbook(str(KGL_LIST_PATH), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore apertura file: {e}")

    if "KG-L" not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=422, detail='Foglio "KG-L" non trovato nel file')

    ws = wb["KG-L"]
    items = []

    # Riga 1: metadato Power Query → skip
    # Riga 2: header (Nr., Descrizione, ..., PESO CORRETTO, ..., KG\L) → skip
    # Dati da riga 3; col A = item_no, col D (indice 3) = peso_corretto, col G (indice 6) = kgl_l
    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        item_no = row[0]
        peso    = row[3]
        if not item_no or peso is None:
            continue
        item_no_str = str(item_no).strip()
        if not item_no_str:
            continue
        try:
            peso_val = float(str(peso).replace(",", "."))
        except (ValueError, TypeError):
            continue
        if peso_val <= 0:
            continue
        kgl_l_val = None
        raw_kgl_l = row[6] if len(row) > 6 else None
        if raw_kgl_l is not None:
            try:
                v = float(str(raw_kgl_l).replace(",", "."))
                if v > 0:
                    kgl_l_val = v
            except (ValueError, TypeError):
                pass
        items.append({"item_no": item_no_str, "peso_corretto": peso_val, "kgl_l": kgl_l_val})

    wb.close()
    return {"items": items, "count": len(items)}



# ── Bridge Sync ──────────────────────────────────────────────────────────────

_BRIDGE_SUBFOLDER = r"15 - Converitore item list NAV\BRIDGE_Converter_FTCHUB.xlsx"


def _extract_item_type_bi(stat_name: str) -> str:
    if stat_name and str(stat_name).strip() == "Fixed":
        return "NA CORE"
    return "TAIL"


def _parse_master_bi(ws) -> list[dict]:
    result = {}
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        barcode_ext = None
        try:
            if row[5] is not None:
                barcode_ext = int(float(str(row[5])))
        except (ValueError, TypeError):
            pass
        result[item_no] = {
            "item_no":      item_no,
            "category":     str(row[3]).strip() if row[3] is not None else None,
            "subcategory":  str(row[4]).strip() if row[4] is not None else None,
            "barcode_ext":  barcode_ext,
            "item_type_bi": _extract_item_type_bi(str(row[2]) if row[2] is not None else ""),
        }
    return list(result.values())


def _parse_price(ws) -> list[dict]:
    items = []
    for row in ws.iter_rows(min_row=5, max_col=2, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        try:
            crp = float(str(row[1]).replace(",", "."))
        except (ValueError, TypeError):
            continue
        items.append({"item_no": item_no, "country_rp": crp})
    return items


_DISPLAY_KEYWORDS = [
    "Table", "Wall", "Behind the till", "Fridge", "Card wall",
    "Surprice bag area", "Bin", "Sales unit", "Candle wall",
]


def _extract_display_modulo(vm_module: str) -> str:
    if not vm_module or not vm_module.strip():
        return "ND"
    for kw in _DISPLAY_KEYWORDS:
        if kw in vm_module:
            return kw
    return "ND"


def _parse_display(ws) -> list[dict]:
    result = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        vm_module = str(row[1]).strip() if row[1] is not None else ""
        try:
            flag = bool(int(float(str(row[2])))) if row[2] is not None else False
        except (ValueError, TypeError):
            flag = False
        result[item_no] = {
            "item_no":              item_no,
            "vm_module":            vm_module or None,
            "flag_hanging_display": flag,
            "modulo":               _extract_display_modulo(vm_module),
        }
    return list(result.values())


def _parse_box_size(ws) -> list[dict]:
    result = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        item_no = str(row[0]).strip() if row[0] is not None else ""
        if not item_no:
            continue
        try:
            box_val = int(float(str(row[2])))
            result[item_no] = max(box_val, 1)
        except (ValueError, TypeError):
            continue
    return [{"item_no": k, "box_size": v} for k, v in result.items()]


BRIDGE_SHEETS = [
    {
        "key":      "price",
        "sheet":    "Price",
        "endpoint": "/api/items/converter/price/import-json",
        "parse":    _parse_price,
    },
    {
        "key":      "box_size",
        "sheet":    "BOX SIZE",
        "endpoint": "/api/items/converter/box-size/import-json",
        "parse":    _parse_box_size,
    },
    {
        "key":      "display",
        "sheet":    "Display",
        "endpoint": "/api/items/converter/display/import-json",
        "parse":    _parse_display,
    },
    {
        "key":      "master_bi",
        "sheet":    "Master Data BI",
        "endpoint": "/api/items/converter/master-bi/import-json",
        "parse":    _parse_master_bi,
    },
]


def _refresh_excel_bridge(file_path: Path) -> None:
    import pythoncom
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(file_path))
        for conn in wb.Connections:
            try:
                conn.OLEDBConnection.BackgroundQuery = False
            except Exception:
                try:
                    conn.ODBCConnection.BackgroundQuery = False
                except Exception:
                    pass
        wb.RefreshAll()
        xl.CalculateUntilAsyncQueriesDone()
        wb.Save()
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        try:
            xl.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


class BridgeSyncRequest(BaseModel):
    no_excel:    bool = False
    backend_url: str
    token:       str


@app.post("/bridge-sync")
def bridge_sync(req: BridgeSyncRequest):
    headers = {"Authorization": f"Bearer {req.token}", "Content-Type": "application/json"}
    base = req.backend_url.rstrip("/")
    ssl = base.startswith("https")

    # Leggi commercial_files_path dal backend
    try:
        r = requests.get(f"{base}/api/items/converter/config", headers=headers, verify=not ssl, timeout=15)
        r.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Errore config backend: {e}")

    commercial_path = r.json().get("commercial_files_path", "").strip()
    if not commercial_path:
        raise HTTPException(status_code=400, detail="commercial_files_path non configurato in FTC HUB — Impostazioni > One Italy Commercial")

    username = os.environ.get("USERNAME", "")
    converter_path = Path(f"C:\\Users\\{username}\\{commercial_path}\\{_BRIDGE_SUBFOLDER}")

    if not converter_path.exists():
        raise HTTPException(status_code=404, detail=f"File non trovato: {converter_path}")

    if not req.no_excel:
        try:
            _refresh_excel_bridge(converter_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Errore refresh Excel: {e}")

    wb = openpyxl.load_workbook(str(converter_path), read_only=True, data_only=True)
    results = {}
    errors  = []

    for cfg in BRIDGE_SHEETS:
        if cfg["sheet"] not in wb.sheetnames:
            errors.append(f'Foglio "{cfg["sheet"]}" non trovato nel file')
            continue
        try:
            items = cfg["parse"](wb[cfg["sheet"]])
            r = requests.post(
                f"{base}{cfg['endpoint']}",
                json={"items": items},
                headers=headers,
                verify=not ssl,
                timeout=180,
            )
            r.raise_for_status()
            results[cfg["key"]] = r.json().get("synced", len(items))
        except Exception as e:
            errors.append(f'{cfg["key"]}: {e}')

    wb.close()
    return {"results": results, "errors": errors}


# ── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=19999, log_level="warning")
